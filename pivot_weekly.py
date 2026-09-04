"""
Weekly Pivot Report — Aug 28–Sep 3 2026  (Week 36)
Generates Excel pivot: cases raised by channel, recovery status, amounts
"""
import sys, os, json
sys.path.insert(0, r'C:\Users\Admin\recon-dashboard')

from sheets_service import get_india_us_data, get_us2us_data
from auth import get_sheets_credentials
from datetime import date, datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter

# ── Config / credentials ──────────────────────────────────────────────────
with open(r'C:\Users\Admin\recon-dashboard\config.json') as f:
    CONFIG = json.load(f)

CF       = CONFIG['credentials_file']
TF       = CONFIG['token_file']
SHEET_ID = CONFIG['recon_sheet_id']   # India→US and US→US both live here

creds = get_sheets_credentials(CF, TF)

# ── Date range ──────────────────────────────────────────────────────────────
START        = date(2026, 8, 28)           # Week 36 — Thu Aug 28
END          = date(2026, 9,  3)           # Week 36 — Wed Sep 3
MID_WK_END   = date(2026, 8, 27)          # W35 end (Aug 27)
LAST_WK_END  = date(2026, 8, 20)          # W34 end (Aug 20)
WEEK_NUM     = START.isocalendar()[1]
MID_WEEK     = MID_WK_END.isocalendar()[1]
PREV_WEEK    = LAST_WK_END.isocalendar()[1]

FOLDER    = r'C:\Users\Admin\Desktop\WeeklyPivotReports'
SNAP_FILE = os.path.join(FOLDER, 'snapshots.json')

# ── Helpers ─────────────────────────────────────────────────────────────────
DATE_FMTS = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y',
             '%Y/%m/%d', '%d %b %Y', '%b %d %Y']

def parse_date(s):
    if not s:
        return None
    s = str(s).strip().split('T')[0]   # strip time part if ISO datetime
    for fmt in DATE_FMTS:
        try:
            return datetime.strptime(s, fmt).date()
        except:
            pass
    return None

def in_range(d):
    return d and START <= d <= END

# ── Fetch data ───────────────────────────────────────────────────────────────
print("Fetching India→US data …")
ius_raw  = get_india_us_data(creds, SHEET_ID)
ius_rows = ius_raw.get('rows', [])

print("Fetching US→US data …")
u2u_raw  = get_us2us_data(creds, SHEET_ID)
u2u_rows = u2u_raw.get('rows', [])

# ── Pivot builders ───────────────────────────────────────────────────────────
# Three views per channel:
#   RAISED_NEW   : raised this week (new filings)  → amt = expected/claimed
#   CLOSED_OLD   : closed this week BUT raised BEFORE this week (backlog clearance)
#   CLOSED_NEW   : raised AND closed this week (same-week turnaround)

def empty():
    return {'labels': set(), 'exp': 0.0, 'rec': 0.0}

def build_pivots(rows, ch_key, raise_key, close_key, amt_key):
    raised_new = defaultdict(empty)   # filed this week
    closed_new = defaultdict(empty)   # raised & closed same week

    # closed_old keyed by (channel, raise_month_label, raise_week_no)
    # value: {'labels': set, 'exp': float, 'rec': float}
    closed_old = defaultdict(lambda: defaultdict(empty))

    seen_rn = set()
    seen_co = set()
    seen_cn = set()

    for r in rows:
        ch  = (r.get(ch_key) or 'Unknown').strip()
        lbl = r.get('label') or r.get('from_label') or r.get('awb') or str(id(r))

        rd  = parse_date(r.get(raise_key))
        cd  = parse_date(r.get(close_key))
        exp = float(r.get(amt_key[0]) or 0)
        rec = float(r.get(amt_key[1]) or 0)

        raised_this_week = in_range(rd)
        closed_this_week = in_range(cd)

        # New filing this week — track Expected only; received shows in closed sections
        if raised_this_week:
            key = (ch, lbl)
            if key not in seen_rn:
                seen_rn.add(key)
                raised_new[ch]['labels'].add(lbl)
                raised_new[ch]['exp'] += exp
                # rec intentionally NOT added here — same-week cases show received in closed_new

            # Same-week turnaround
            if closed_this_week:
                if key not in seen_cn:
                    seen_cn.add(key)
                    closed_new[ch]['labels'].add(lbl)
                    closed_new[ch]['exp'] += exp
                    closed_new[ch]['rec'] += rec

        # Old case closed this week — group by when it was originally raised
        elif closed_this_week:
            key = (ch, lbl)
            if key not in seen_co:
                seen_co.add(key)
                if rd:
                    wk  = rd.isocalendar()[1]
                    mon = rd.strftime('%b %Y')          # e.g. "Jun 2026"
                    age_label = f"{mon}  (W{wk})"
                else:
                    age_label = 'Unknown raise date'
                closed_old[ch][age_label]['labels'].add(lbl)
                closed_old[ch][age_label]['exp'] += exp
                closed_old[ch][age_label]['rec'] += rec

    return raised_new, closed_old, closed_new

# India→US
ius_raised_new, ius_closed_old, ius_closed_new = build_pivots(
    ius_rows,
    ch_key    = 'channel',
    raise_key = 'case_raise_date',
    close_key = 'case_close_date',
    amt_key   = ('expected', 'actual'),
)

# US→US
u2u_raised_new, u2u_closed_old, u2u_closed_new = build_pivots(
    u2u_rows,
    ch_key    = 'to_channel',
    raise_key = 'case_raise_date',
    close_key = 'case_resolve_date',
    amt_key   = ('expected', 'recovered'),
)

# ── All-time totals per channel (no date filter) ──────────────────────────────
def build_alltime(rows, ch_key, raise_key, amt_key):
    """All cases ever raised, grouped by channel."""
    totals = defaultdict(lambda: {'labels': set(), 'claimed': 0.0, 'recovered': 0.0})
    seen   = set()
    for r in rows:
        ch  = (r.get(ch_key) or 'Unknown').strip()
        lbl = r.get('label') or r.get('from_label') or r.get('awb') or str(id(r))
        rd  = parse_date(r.get(raise_key))
        if not rd:
            continue            # skip rows with no raise date (not yet filed)
        key = (ch, lbl)
        if key not in seen:
            seen.add(key)
            totals[ch]['labels'].add(lbl)
            totals[ch]['claimed']   += float(r.get(amt_key[0]) or 0)
            totals[ch]['recovered'] += float(r.get(amt_key[1]) or 0)
    return totals

ius_alltime = build_alltime(ius_rows, 'channel',    'case_raise_date', ('expected', 'actual'))
u2u_alltime = build_alltime(u2u_rows, 'to_channel', 'case_raise_date', ('expected', 'recovered'))

# ── Claims Bucket snapshot — as of a given cutoff date ───────────────────────
# Mirrors the dashboard's "Claims Bucket (YTD)" table logic.
# Three buckets per case:
#   Raised & Received   : case_raise_date ≤ cutoff AND case_close_date ≤ cutoff
#   Raised, Pending     : case_raise_date ≤ cutoff AND NOT yet closed by cutoff
#   Pending to Claim    : no raise date (or raise_date > cutoff), but has expected value
#
# Returns dict: { 'received': float, 'pending_receipt': float, 'pending_claim': float }

RAISED_KEYWORDS = [
    'claim raised but not received', 'claim raised but pending',
    'claim submitted', 'claim filed', 'case raised',
    'under review', 'claim in progress', 'raised',
]
PENDING_KEYWORDS = ['pending', 'to be claimed']

def claims_snapshot(rows, cutoff, close_key, exp_key, act_key):
    """
    Time-aware Claims Bucket snapshot as of a given cutoff date.
      received        : actual > 0  AND  close_date ≤ cutoff
      pending_receipt : case_raise_date ≤ cutoff AND not yet paid  (time-aware)
      pending_claim   : sub_remark == 'Pending to claim' exactly  (no date anchor)
    """
    buckets = {'received': 0.0, 'pending_receipt': 0.0, 'pending_claim': 0.0}
    for r in rows:
        cd  = parse_date(r.get(close_key))
        exp = float(r.get(exp_key) or 0)
        act = float(r.get(act_key) or 0)

        if exp == 0 and act == 0:
            continue

        closed_by_cutoff = (not cd) or (cd <= cutoff)

        sl = ((r.get('sub_remark') or '') + ' ' + (r.get('remarks') or '')).lower().strip()

        # Use only explicit case_raise_date; check it is ≤ cutoff for time-awareness
        explicit_rd = parse_date(r.get('case_raise_date'))
        raised_by_cutoff = bool(explicit_rd) and explicit_rd <= cutoff

        if act > 0 and closed_by_cutoff:
            buckets['received'] += act
        elif act > 0 and not closed_by_cutoff:
            # paid after cutoff → was pending at that point in time
            buckets['pending_receipt'] += exp
        elif act == 0 and exp > 0:
            if raised_by_cutoff or any(k in sl for k in RAISED_KEYWORDS):
                # Claim was raised on/before cutoff but not yet paid
                buckets['pending_receipt'] += exp
            elif (r.get('sub_remark') or '').strip().lower() == 'pending to claim':
                buckets['pending_claim'] += exp
            # else: IBR/WRO or other — skip

    return {k: round(v, 2) for k, v in buckets.items()}

# India→US snapshots — computed from live data
ius_snap_prev = claims_snapshot(ius_rows, LAST_WK_END, 'case_close_date',   'expected', 'actual')
ius_snap_mid  = claims_snapshot(ius_rows, MID_WK_END,  'case_close_date',   'expected', 'actual')
ius_snap_curr = claims_snapshot(ius_rows, END,          'case_close_date',   'expected', 'actual')

# US→US snapshots — computed from live data
u2u_snap_prev = claims_snapshot(u2u_rows, LAST_WK_END, 'case_resolve_date', 'expected', 'recovered')
u2u_snap_mid  = claims_snapshot(u2u_rows, MID_WK_END,  'case_resolve_date', 'expected', 'recovered')
u2u_snap_curr = claims_snapshot(u2u_rows, END,          'case_resolve_date', 'expected', 'recovered')

print(f"\nClaims Bucket (sub_remark='Pending to claim' only):")
print(f"  India→US  {END.strftime('%b %d')} → received: ${ius_snap_curr['received']:,.2f}  "
      f"| pending_receipt: ${ius_snap_curr['pending_receipt']:,.2f}  "
      f"| pending_claim: ${ius_snap_curr['pending_claim']:,.2f}")
print(f"  US→US     {END.strftime('%b %d')} → received: ${u2u_snap_curr['received']:,.2f}  "
      f"| pending_receipt: ${u2u_snap_curr['pending_receipt']:,.2f}  "
      f"| pending_claim: ${u2u_snap_curr['pending_claim']:,.2f}")

# ── Auto-load reference snapshots from previous runs ─────────────────────────
# snapshots.json is written after each run; Ref* column fills itself each week.
os.makedirs(FOLDER, exist_ok=True)
_all_snaps = {}
if os.path.exists(SNAP_FILE):
    with open(SNAP_FILE) as _f:
        _all_snaps = json.load(_f)

def _get_ref(dt, route):
    entry = _all_snaps.get(dt.isoformat(), {})
    return entry.get(route)   # None if missing → shows '—' in Ref column

REF = {
    'ius': {
        LAST_WK_END: _get_ref(LAST_WK_END, 'ius'),
        MID_WK_END:  _get_ref(MID_WK_END,  'ius'),
        END:         None,
    },
    'u2u': {
        LAST_WK_END: _get_ref(LAST_WK_END, 'u2u'),
        MID_WK_END:  _get_ref(MID_WK_END,  'u2u'),
        END:         None,
    }
}
print(f"[Snapshots] Loaded {len(_all_snaps)} dates from snapshots.json")
for _rt in ('ius', 'u2u'):
    for _dt, _v in REF[_rt].items():
        if _dt == END:
            continue
        tag = _v and f"received=${_v['received']:,.2f}" or '— (no snapshot)'
        print(f"  REF {_rt} {_dt}: {tag}")

# ── Debug summary ─────────────────────────────────────────────────────────────
def show(name, raised_new, closed_old, closed_new):
    print(f"\n{'='*72}")
    print(f"{name}  |  Week {WEEK_NUM}  |  {START} → {END}")
    print(f"{'='*72}")
    all_ch = sorted(set(
        list(raised_new.keys()) + list(closed_old.keys()) + list(closed_new.keys())
    ))
    if not all_ch:
        print("  (no activity in date range)")
        return
    for ch in all_ch:
        rn = raised_new[ch]
        cn = closed_new[ch]
        co_by_age = closed_old.get(ch, {})
        co_total  = sum(len(v['labels']) for v in co_by_age.values())
        co_rec    = sum(v['rec'] for v in co_by_age.values())
        print(f"  {ch}")
        print(f"    Raised new   : {len(rn['labels']):>3} cases  ${rn['exp']:>9.2f} expected  ${rn['rec']:>9.2f} received")
        print(f"    Closed (old) : {co_total:>3} cases  ${co_rec:>9.2f} received  ← from backlog")
        for age_lbl in sorted(co_by_age.keys()):
            v = co_by_age[age_lbl]
            print(f"         {age_lbl:<30} {len(v['labels']):>2} cases  ${v['rec']:>8.2f} received")
        print(f"    Closed (new) : {len(cn['labels']):>3} cases  ${cn['exp']:>9.2f} expected  ${cn['rec']:>9.2f} received  ← same-week")

show("India → US", ius_raised_new, ius_closed_old, ius_closed_new)
show("US → US",    u2u_raised_new, u2u_closed_old, u2u_closed_new)

# ── Cross-check against direct sheet filter ───────────────────────────────────
print(f"\n{'='*72}")
print(f"Cross-check  |  filter case_close_date in [{START} – {END}]")
ius_cls_chk = sum(float(r.get('actual') or 0)
                  for r in ius_rows if in_range(parse_date(r.get('case_close_date'))))
u2u_cls_chk = sum(float(r.get('recovered') or 0)
                  for r in u2u_rows if in_range(parse_date(r.get('case_resolve_date'))))
print(f"  India→US  actual received : ${ius_cls_chk:>9.2f}")
print(f"  US→US     actual received : ${u2u_cls_chk:>9.2f}")
print(f"  Combined                  : ${ius_cls_chk + u2u_cls_chk:>9.2f}")

print(f"\nCross-check  |  filter case_raise_date in [{START} – {END}]")
ius_raise_rows = [r for r in ius_rows if in_range(parse_date(r.get('case_raise_date')))]
u2u_raise_rows = [r for r in u2u_rows if in_range(parse_date(r.get('case_raise_date')))]
ius_exp_chk = sum(float(r.get('expected') or 0) for r in ius_raise_rows)
u2u_exp_chk = sum(float(r.get('expected') or 0) for r in u2u_raise_rows)
print(f"  India→US  raised: {len(ius_raise_rows):>3} rows   expected ${ius_exp_chk:>9.2f}")
print(f"  US→US     raised: {len(u2u_raise_rows):>3} rows   expected ${u2u_exp_chk:>9.2f}")
print(f"{'='*72}")

# ── Excel builder ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default sheet

# Styles
HDR_BG   = PatternFill('solid', fgColor='1A237E')   # deep navy
IUS_BG   = PatternFill('solid', fgColor='E3F2FD')   # light blue for IU→US rows
U2U_BG   = PatternFill('solid', fgColor='E8F5E9')   # light green for US→US rows
TOT_BG   = PatternFill('solid', fgColor='FFF3E0')   # amber for totals
CLOSE_BG = PatternFill('solid', fgColor='C8E6C9')   # green for closed count
OPEN_BG  = PatternFill('solid', fgColor='FFCCBC')   # orange for open count

HDR_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
TITLE_FONT= Font(name='Calibri', bold=True, color='1A237E', size=12)
DATA_FONT = Font(name='Calibri', size=10)
BOLD_FONT = Font(name='Calibri', bold=True, size=10)
TOT_FONT  = Font(name='Calibri', bold=True, color='BF360C', size=10)

THIN  = Side(style='thin',   color='B0BEC5')
MED   = Side(style='medium', color='90A4AE')
def bdr(t=None, l=None, r=None, b=None):
    return Border(top=t or THIN, left=l or THIN, right=r or THIN, bottom=b or THIN)

USD_FMT  = '"$"#,##0.00'
PCT_FMT  = '0.0"%"'
NUM_FMT  = '#,##0'

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=False)
LEFT   = Alignment(horizontal='left',   vertical='center')
RIGHT  = Alignment(horizontal='right',  vertical='center')

def style(cell, fill=None, font=None, align=None, fmt=None, bdr_=None):
    if fill:  cell.fill      = fill
    if font:  cell.font      = font
    if align: cell.alignment = align
    if fmt:   cell.number_format = fmt
    if bdr_:  cell.border    = bdr_

SECT_BG_RAISED = PatternFill('solid', fgColor='E3F2FD')   # blue tint — raised section header
SECT_BG_CLOSED = PatternFill('solid', fgColor='E8F5E9')   # green tint — closed section header
SECT_FONT      = Font(name='Calibri', bold=True, color='37474F', size=10)

def write_section(ws, row_num, section_title, sect_fill, col_headers,
                  data_rows, row_fill, col_fmts, col_aligns):
    """Write a labelled section (header + column headers + data + totals)."""
    NCOLS = len(col_headers)

    # Section label
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=NCOLS)
    c = ws.cell(row=row_num, column=1, value=section_title)
    style(c, fill=sect_fill, font=SECT_FONT, align=LEFT)
    ws.row_dimensions[row_num].height = 18
    row_num += 1

    # Column headers
    for col, h in enumerate(col_headers, 1):
        c = ws.cell(row=row_num, column=col, value=h)
        style(c, fill=HDR_BG, font=HDR_FONT, align=CENTER)
    ws.row_dimensions[row_num].height = 17
    row_num += 1

    if not data_rows:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=NCOLS)
        c = ws.cell(row=row_num, column=1, value='No activity in this date range')
        style(c, font=Font(name='Calibri', italic=True, color='9E9E9E', size=10), align=CENTER)
        ws.row_dimensions[row_num].height = 16
        return row_num + 2

    totals = [0] * NCOLS
    for vals in data_rows:
        for col, (v, fm, al) in enumerate(zip(vals, col_fmts, col_aligns), 1):
            c = ws.cell(row=row_num, column=col, value=v)
            style(c, fill=(row_fill if col > 1 else None), font=DATA_FONT, align=al, fmt=fm, bdr_=bdr())
            if isinstance(v, (int, float)):
                totals[col-1] += v
        ws.row_dimensions[row_num].height = 17
        row_num += 1

    # Totals row
    tot_vals = ['TOTAL'] + totals[1:]
    for col, (v, fm) in enumerate(zip(tot_vals, col_fmts), 1):
        c = ws.cell(row=row_num, column=col, value=v)
        style(c, fill=TOT_BG, font=TOT_FONT,
              align=(LEFT if col==1 else (CENTER if isinstance(v, int) else RIGHT)),
              fmt=fm, bdr_=bdr(t=Side(style='medium', color='BF360C')))
    ws.row_dimensions[row_num].height = 18
    return row_num + 2   # gap after section


SECT_BG_SAMEWEEK = PatternFill('solid', fgColor='FFF9C4')   # yellow — same-week turnaround

def add_pivot_sheet(wb, title, raised_new, closed_old, closed_new, tab_label):
    ws = wb.create_sheet(title=title)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"📊  {tab_label}  —  Week {WEEK_NUM}  ({START.strftime('%d %b')} – {END.strftime('%d %b %Y')})"
    style(ws['A1'], font=TITLE_FONT, align=LEFT)
    ws.row_dimensions[1].height = 22

    ws.merge_cells('A2:F2')
    ws['A2'] = ('Section 1: new cases raised this week  |  '
                'Section 2: old backlog cases closed this week  |  '
                'Section 3: raised & resolved same week')
    style(ws['A2'], font=Font(name='Calibri', italic=True, color='546E7A', size=9), align=LEFT)
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 6

    row = 4

    # Section 1 — Raised this week (Expected only; received tracked in closed sections)
    rn_rows = [[ch, len(m['labels']), round(m['exp'],2)]
               for ch, m in sorted(raised_new.items())]
    row = write_section(ws, row,
        '📋  Cases RAISED this week  (new filings — Case Raise Date in range)',
        SECT_BG_RAISED,
        ['Channel', 'Cases Raised', 'Expected ($)'],
        rn_rows, IUS_BG,
        [None, NUM_FMT, USD_FMT],
        [LEFT, CENTER, RIGHT])

    # Section 2 — Old cases closed this week, broken down by when they were raised
    co_rows = []
    for ch in sorted(closed_old.keys()):
        by_age = closed_old[ch]
        ch_cnt = sum(len(v['labels']) for v in by_age.values())
        ch_exp = round(sum(v['exp'] for v in by_age.values()), 2)
        ch_rec = round(sum(v['rec'] for v in by_age.values()), 2)
        # Channel summary row
        co_rows.append([ch, ch_cnt, ch_exp, ch_rec, '← total for channel'])
        # One sub-row per raise-month/week
        for age_lbl in sorted(by_age.keys()):
            v = by_age[age_lbl]
            co_rows.append(['    ↳ raised in ' + age_lbl, len(v['labels']),
                            round(v['exp'],2), round(v['rec'],2), ''])

    row = write_section(ws, row,
        '💰  Old cases CLOSED this week  (backlog recovery — raised before this week)',
        SECT_BG_CLOSED,
        ['Channel / Raise Period', 'Cases Closed', 'Expected ($)', 'Received ($)', 'Note'],
        co_rows, U2U_BG,
        [None, NUM_FMT, USD_FMT, USD_FMT, None],
        [LEFT, CENTER, RIGHT, RIGHT, LEFT])

    # Section 3 — Same-week turnaround
    cn_rows = [[ch, len(m['labels']), round(m['exp'],2), round(m['rec'],2)]
               for ch, m in sorted(closed_new.items())]
    row = write_section(ws, row,
        '⚡  Raised & closed SAME WEEK  (fast turnaround — raised and resolved this week)',
        SECT_BG_SAMEWEEK,
        ['Channel', 'Cases', 'Expected ($)', 'Received ($)'],
        cn_rows, PatternFill('solid', fgColor='FFF9C4'),
        [None, NUM_FMT, USD_FMT, USD_FMT], [LEFT, CENTER, RIGHT, RIGHT])

    for i, w in enumerate([34, 12, 16, 16, 16, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'
    return ws


# Create sheets
add_pivot_sheet(wb, 'India → US', ius_raised_new, ius_closed_old, ius_closed_new, 'India → US Claims')
add_pivot_sheet(wb, 'US → US',    u2u_raised_new, u2u_closed_old, u2u_closed_new, 'US → US Claims')

# ── Combined summary sheet ────────────────────────────────────────────────────
ws_sum = wb.create_sheet(title='Summary', index=0)
ws_sum.sheet_view.showGridLines = False

ws_sum.merge_cells('A1:H1')
ws_sum['A1'] = f"📋  Weekly Recovery Summary  —  Week {WEEK_NUM}  ({START.strftime('%d %b')} – {END.strftime('%d %b %Y')})"
style(ws_sum['A1'], font=TITLE_FONT, align=LEFT)
ws_sum.row_dimensions[1].height = 24

ws_sum.merge_cells('A2:H2')
ws_sum['A2'] = 'RAISED = cases filed this week  |  CLOSED = cases resolved this week  |  Expected = amount claimed  |  Received = amount paid back'
style(ws_sum['A2'], font=Font(name='Calibri', italic=True, color='546E7A', size=9), align=LEFT)
ws_sum.row_dimensions[2].height = 14
ws_sum.row_dimensions[3].height = 6

sum_hdrs = ['Route', 'View', 'Channel', 'Case Count', 'Expected ($)', 'Received ($)', 'Note']
for col, h in enumerate(sum_hdrs, 1):
    c = ws_sum.cell(row=4, column=col, value=h)
    style(c, fill=HDR_BG, font=HDR_FONT, align=CENTER)
ws_sum.row_dimensions[4].height = 18

sr = 5
def write_sum_section(route, view_label, pivot_dict, fill, show_rec=True):
    global sr
    for ch in sorted(pivot_dict.keys()):
        m = pivot_dict[ch]
        cnt = len(m['labels'])
        exp = round(m['exp'], 2)
        rec = round(m['rec'], 2) if show_rec else '—'
        rec_fmt = USD_FMT if show_rec else None
        note = '' if show_rec else 'Awaiting recovery'
        for col, (v, fm, al) in enumerate(zip(
            [route, view_label, ch, cnt, exp, rec, note],
            [None, None, None, NUM_FMT, USD_FMT, rec_fmt, None],
            [LEFT, CENTER, LEFT, CENTER, RIGHT, RIGHT, LEFT]
        ), 1):
            c = ws_sum.cell(row=sr, column=col, value=v)
            style(c, fill=fill, font=DATA_FONT, align=al, fmt=fm, bdr_=bdr())
        ws_sum.row_dimensions[sr].height = 17
        sr += 1

def write_old_sum(route, pivot_nested, fill):
    """Write one row per (channel + raise week) for closed_old backlog."""
    global sr
    for ch in sorted(pivot_nested.keys()):
        by_age = pivot_nested[ch]
        for age_lbl in sorted(by_age.keys()):
            v = by_age[age_lbl]
            cnt = len(v['labels'])
            exp = round(v['exp'], 2)
            rec = round(v['rec'], 2)
            view_label = f'💰 Closed — old backlog  (raised {age_lbl})'
            for col, (val, fm, al) in enumerate(zip(
                [route, view_label, ch, cnt, exp, rec, ''],
                [None, None, None, NUM_FMT, USD_FMT, USD_FMT, None],
                [LEFT, LEFT, LEFT, CENTER, RIGHT, RIGHT, LEFT]
            ), 1):
                c = ws_sum.cell(row=sr, column=col, value=val)
                style(c, fill=fill, font=DATA_FONT, align=al, fmt=fm, bdr_=bdr())
            ws_sum.row_dimensions[sr].height = 17
            sr += 1

write_sum_section('India → US', '📋 RAISED (new)',       ius_raised_new, IUS_BG,                                  show_rec=False)
write_old_sum    ('India → US',                         ius_closed_old, PatternFill('solid', fgColor='C8E6C9'))
write_sum_section('India → US', '⚡ CLOSED (same week)', ius_closed_new, PatternFill('solid', fgColor='FFF9C4'),  show_rec=True)
write_sum_section('US → US',    '📋 RAISED (new)',       u2u_raised_new, U2U_BG,                                  show_rec=False)
write_old_sum    ('US → US',                             u2u_closed_old, PatternFill('solid', fgColor='A5D6A7'))
write_sum_section('US → US',    '⚡ CLOSED (same week)',  u2u_closed_new, PatternFill('solid', fgColor='FFF9C4'),  show_rec=True)

# ── All-Time section ─────────────────────────────────────────────────────────
sr += 1   # blank gap row
ALLTIME_HDR_BG = PatternFill('solid', fgColor='37474F')
ALLTIME_ROW_BG = PatternFill('solid', fgColor='ECEFF1')

ws_sum.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=7)
c = ws_sum.cell(row=sr, column=1,
    value='📊  All-Time Cases Raised by Channel  (all weeks combined, cases with a raise date)')
style(c, fill=PatternFill('solid', fgColor='263238'),
      font=Font(name='Calibri', bold=True, color='FFFFFF', size=10), align=LEFT)
ws_sum.row_dimensions[sr].height = 18
sr += 1

at_hdrs = ['Route', 'Channel', 'Total Raised (all time)',
           'This Week Raised', 'This Week %', 'Total Claimed ($)']
for col, h in enumerate(at_hdrs, 1):
    c = ws_sum.cell(row=sr, column=col, value=h)
    style(c, fill=ALLTIME_HDR_BG, font=HDR_FONT, align=CENTER)
ws_sum.row_dimensions[sr].height = 17
sr += 1

def write_alltime_rows(route, alltime, raised_this_week):
    global sr
    for ch in sorted(alltime.keys()):
        m       = alltime[ch]
        total   = len(m['labels'])
        claimed = round(m['claimed'], 2)
        this_wk = len(raised_this_week.get(ch, empty())['labels'])
        pct     = round(this_wk / total * 100, 1) if total > 0 else 0.0
        for col, (v, fm, al) in enumerate(zip(
            [route, ch, total, this_wk, pct, claimed],
            [None, None, NUM_FMT, NUM_FMT, PCT_FMT, USD_FMT],
            [LEFT, LEFT, CENTER, CENTER, CENTER, RIGHT]
        ), 1):
            c = ws_sum.cell(row=sr, column=col, value=v)
            style(c, fill=ALLTIME_ROW_BG, font=DATA_FONT, align=al, fmt=fm, bdr_=bdr())
        ws_sum.row_dimensions[sr].height = 17
        sr += 1

write_alltime_rows('India → US', ius_alltime, ius_raised_new)
write_alltime_rows('US → US',    u2u_alltime, u2u_raised_new)

sum_widths = [14, 36, 22, 14, 16, 16, 20]
for i, w in enumerate(sum_widths, 1):
    ws_sum.column_dimensions[get_column_letter(i)].width = w
ws_sum.freeze_panes = 'A5'

# ── WoW Claims Bucket sheet ───────────────────────────────────────────────────
ws_wow = wb.create_sheet(title='WoW Claims Bucket')
ws_wow.sheet_view.showGridLines = False

BUCKET_LABELS = [
    ('received',        '✅ Claim Raised & Received',      '2e7d32', 'C8E6C9'),
    ('pending_receipt', '⏳ Claim Raised, Pending Receipt', 'e65100', 'FFF3E0'),
    ('pending_claim',   '📋 Pending to Claim',             '1565c0', 'E3F2FD'),
]

UP_COLOR   = '2e7d32'   # green — went up (more recovered)
DOWN_COLOR = 'c62828'   # red   — went down
NEU_COLOR  = '546E7A'   # grey  — no change

def wow_sheet(ws, route_label, snap_prev, snap_mid, snap_curr, ref_prev, ref_mid):
    """
    3-column WoW comparison: Aug 6 | Aug 13 | Aug 20
    Each date shows: Total ($) | YTD% | Ref total if available
    Final two columns: WoW change (Aug 13→20) and full change (Aug 6→20)
    """
    row = 1

    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1,
            value=f"📊  Claims Bucket — 3-Week Comparison  |  {route_label}")
    style(ws.cell(row=row, column=1), font=TITLE_FONT, align=LEFT)
    ws.row_dimensions[row].height = 22
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1,
            value=(f"{LAST_WK_END.strftime('%d %b')} (W{PREV_WEEK})  →  "
                   f"{MID_WK_END.strftime('%d %b')} (W{MID_WEEK})  →  "
                   f"{END.strftime('%d %b %Y')} (W{WEEK_NUM} current)  "
                   f"|  * Ref = dashboard screenshot value"))
    style(ws.cell(row=row, column=1),
          font=Font(name='Calibri', italic=True, color='546E7A', size=9), align=LEFT)
    ws.row_dimensions[row].height = 14
    row += 2

    # ── Column headers ──────────────────────────────────────────────────────
    hdrs = [
        'Claims Bucket',
        f'{LAST_WK_END.strftime("%d %b")} Total',
        f'{LAST_WK_END.strftime("%d %b")} YTD%',
        f'{LAST_WK_END.strftime("%d %b")} Ref *',
        f'{MID_WK_END.strftime("%d %b")} Total',
        f'{MID_WK_END.strftime("%d %b")} YTD%',
        f'{MID_WK_END.strftime("%d %b")} Ref *',
        f'{END.strftime("%d %b")} Total',
        f'{END.strftime("%d %b")} YTD%',
        f'WoW Change ({MID_WK_END.strftime("%d")}-{END.strftime("%d %b")})',
    ]
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=col, value=h)
        style(c, fill=HDR_BG, font=HDR_FONT, align=CENTER)
    ws.row_dimensions[row].height = 22
    row += 1

    prev_total = sum(snap_prev.values())
    mid_total  = sum(snap_mid.values())
    curr_total = sum(snap_curr.values())

    ref_prev_total = sum(ref_prev.values()) if ref_prev else None
    ref_mid_total  = sum(ref_mid.values())  if ref_mid  else None

    for bkey, blabel, bcolor, bfill_hex in BUCKET_LABELS:
        pv  = snap_prev.get(bkey, 0.0)
        mv  = snap_mid.get(bkey, 0.0)
        cv  = snap_curr.get(bkey, 0.0)

        ppct = round(pv / prev_total * 100, 1) if prev_total else 0.0
        mpct = round(mv / mid_total  * 100, 1) if mid_total  else 0.0
        cpct = round(cv / curr_total * 100, 1) if curr_total else 0.0

        rp   = ref_prev.get(bkey, 0.0) if ref_prev else None
        rm   = ref_mid.get(bkey,  0.0) if ref_mid  else None

        wow_chg   = round(cv - mv, 2)
        arrow_wow = '▲ ' if wow_chg > 0 else ('▼ ' if wow_chg < 0 else '— ')
        clr_wow   = UP_COLOR if wow_chg > 0 else (DOWN_COLOR if wow_chg < 0 else NEU_COLOR)

        row_fill = PatternFill('solid', fgColor=bfill_hex)

        vals = [
            blabel,
            pv,   ppct,  (rp   if rp   is not None else '—'),
            mv,   mpct,  (rm   if rm   is not None else '—'),
            cv,   cpct,
            arrow_wow + f"${abs(wow_chg):,.2f}",
        ]
        fmts   = [None, USD_FMT, PCT_FMT, USD_FMT,
                        USD_FMT, PCT_FMT, USD_FMT,
                        USD_FMT, PCT_FMT,
                        None]
        aligns = [LEFT, RIGHT, CENTER, RIGHT,
                        RIGHT, CENTER, RIGHT,
                        RIGHT, CENTER, CENTER]

        for col, (v, fm, al) in enumerate(zip(vals, fmts, aligns), 1):
            is_ref = col in (4, 7)
            is_chg = col == 10
            clr = (clr_wow if is_chg else ('666666' if is_ref else bcolor if col == 1 else '212121'))
            fnt = Font(name='Calibri', size=10 if not is_ref else 9,
                       bold=(col == 1),
                       italic=is_ref,
                       color=clr)
            fill = PatternFill('solid', fgColor='F5F5F5') if is_ref else row_fill
            c = ws.cell(row=row, column=col, value=v)
            style(c, fill=fill, font=fnt, align=al,
                  fmt=(fm if not is_ref or isinstance(v, float) else None), bdr_=bdr())
        ws.row_dimensions[row].height = 19
        row += 1

    # ── Totals row ─────────────────────────────────────────────────────────
    pt = round(prev_total, 2)
    mt = round(mid_total,  2)
    ct = round(curr_total, 2)
    wow_t = round(ct - mt, 2)
    arrow_t = '▲ ' if wow_t > 0 else ('▼ ' if wow_t < 0 else '— ')
    clr_t   = UP_COLOR if wow_t > 0 else (DOWN_COLOR if wow_t < 0 else NEU_COLOR)
    rpt = round(ref_prev_total, 2) if ref_prev_total else '—'
    rmt = round(ref_mid_total,  2) if ref_mid_total  else '—'

    tot_vals = ['Total Eligible Claims',
                pt,   100.0, rpt,
                mt,   100.0, rmt,
                ct,   100.0,
                arrow_t + f"${abs(wow_t):,.2f}"]
    tot_fmts = [None, USD_FMT, PCT_FMT, USD_FMT,
                      USD_FMT, PCT_FMT, USD_FMT,
                      USD_FMT, PCT_FMT, None]
    for col, (v, fm) in enumerate(zip(tot_vals, tot_fmts), 1):
        is_chg = col == 10
        clr = (clr_t if is_chg else 'FFFFFF')
        c = ws.cell(row=row, column=col, value=v)
        style(c, fill=PatternFill('solid', fgColor='1B3A2D'),
              font=Font(name='Calibri', bold=True, color=clr, size=10),
              align=(LEFT if col == 1 else (CENTER if col in (3,6,9) else RIGHT)),
              fmt=(fm if isinstance(v, float) else None),
              bdr_=bdr(t=Side(style='medium', color='FFFFFF')))
    ws.row_dimensions[row].height = 20

    ws.column_dimensions['A'].width = 30
    for i, w in enumerate([14, 10, 14, 14, 10, 14, 14, 10, 18], 2):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'


# ── Create WoW sheets ─────────────────────────────────────────────────────────
wow_sheet(ws_wow, 'India → US',
          ius_snap_prev, ius_snap_mid, ius_snap_curr,
          REF['ius'][LAST_WK_END], REF['ius'][MID_WK_END])

ws_wow2 = wb.create_sheet(title='WoW Claims Bucket US→US')
ws_wow2.sheet_view.showGridLines = False
wow_sheet(ws_wow2, 'US → US',
          u2u_snap_prev, u2u_snap_mid, u2u_snap_curr,
          REF['u2u'][LAST_WK_END], REF['u2u'][MID_WK_END])

# ── Raw Dump tab — all case rows for this week's activity ─────────────────────
ws_dump = wb.create_sheet(title='This Week — Raw Log')
ws_dump.sheet_view.showGridLines = False

# Title
ws_dump.merge_cells('A1:L1')
ws_dump['A1'] = f"📋  This Week Activity Log  —  W{WEEK_NUM}  ({START.strftime('%d %b')} – {END.strftime('%d %b %Y')})"
style(ws_dump['A1'], font=TITLE_FONT, align=LEFT)
ws_dump.row_dimensions[1].height = 22

ws_dump.merge_cells('A2:L2')
ws_dump['A2'] = ('Rows from India→US and US→US where case_raise_date or case_close_date falls in this week. '
                 'Bucket = status of each case right now.')
style(ws_dump['A2'], font=Font(name='Calibri', italic=True, color='546E7A', size=9), align=LEFT)
ws_dump.row_dimensions[2].height = 14
ws_dump.row_dimensions[3].height = 6

dump_hdrs = ['Route', 'Bucket', 'Channel', 'Label / AWB',
             'Month', 'Case Raise Date', 'Case Close Date',
             'Expected ($)', 'Actual ($)', 'Sub Remark', 'Remarks', 'Activity This Week']
for col, h in enumerate(dump_hdrs, 1):
    c = ws_dump.cell(row=4, column=col, value=h)
    style(c, fill=HDR_BG, font=HDR_FONT, align=CENTER)
ws_dump.row_dimensions[4].height = 18

BUCKET_FILL = {
    'Raised & Received':   PatternFill('solid', fgColor='C8E6C9'),
    'Raised, Pending':     PatternFill('solid', fgColor='FFF3E0'),
    'Pending to Claim':    PatternFill('solid', fgColor='E3F2FD'),
}
BUCKET_COLOR = {
    'Raised & Received':  '2e7d32',
    'Raised, Pending':    'e65100',
    'Pending to Claim':   '1565c0',
}

def dump_rows(route, rows, close_key, exp_key, act_key, ch_key, raise_key,
              include_all_pending=True):
    """
    week_active : rows where raise OR close happened this week → Raised & Received + Raised Pending
    pending_all : ALL rows that are currently Pending to Claim (no raise date or claim not filed)
    """
    week_active = []
    pending_all = []

    sl_cache = {}
    for r in rows:
        rd  = parse_date(r.get(raise_key))
        cd  = parse_date(r.get(close_key))
        exp = float(r.get(exp_key) or 0)
        act = float(r.get(act_key) or 0)
        if exp == 0 and act == 0:
            continue

        raised_this_wk = in_range(rd)
        closed_this_wk = in_range(cd)
        sl = ((r.get('sub_remark') or '') + ' ' + (r.get('remarks') or '')).lower()

        # Current bucket classification
        # If case_raise_date is filled → claim has been raised regardless of keywords
        has_raise_date = bool(parse_date(r.get(raise_key)))
        sub_lower = (r.get('sub_remark') or '').strip().lower()
        if act > 0:
            bucket = 'Raised & Received'
        elif has_raise_date or any(k in sl for k in RAISED_KEYWORDS):
            bucket = 'Raised, Pending'
        elif sub_lower == 'pending to claim':
            bucket = 'Pending to Claim'
        else:
            continue  # IBR/WRO or other status — exclude from report

        # US→US: use to_label (destination) as the case identifier
        lbl = (r.get('to_label') or r.get('label') or r.get('from_label') or r.get('awb') or ''
               if route == 'US → US'
               else r.get('label') or r.get('awb') or '')

        base = {
            'route':    route,
            'bucket':   bucket,
            'channel':  r.get(ch_key) or '',
            'label':    lbl,
            'month':    r.get('month') or '',
            'raise_dt': r.get(raise_key) or '',
            'close_dt': r.get(close_key) or '',
            'exp':      exp,
            'act':      act,
            'sub':      r.get('sub_remark') or '',
            'remarks':  r.get('remarks') or '',
        }

        # Week activity rows (Raised & Received + Raised Pending touched this week)
        if (raised_this_wk or closed_this_wk) and bucket != 'Pending to Claim':
            activity = []
            if raised_this_wk: activity.append(f"Raised {rd.strftime('%d %b')}")
            if closed_this_wk: activity.append(f"Closed {cd.strftime('%d %b')}")
            week_active.append({**base, 'activity': ' + '.join(activity)})

        # Full pending log — all cases not yet claimed
        if include_all_pending and bucket == 'Pending to Claim':
            pending_all.append({**base, 'activity': 'Not yet claimed'})

    return week_active, pending_all

ius_week, ius_pend = dump_rows('India → US', ius_rows,
    'case_close_date', 'expected', 'actual', 'channel', 'case_raise_date')
u2u_week, u2u_pend = dump_rows('US → US', u2u_rows,
    'case_resolve_date', 'expected', 'recovered', 'to_channel', 'case_raise_date')

BUCKET_ORDER = {'Raised & Received': 0, 'Raised, Pending': 1, 'Pending to Claim': 2}

week_rows = ius_week + u2u_week
week_rows.sort(key=lambda x: (BUCKET_ORDER.get(x['bucket'], 9), x['route'], x['channel']))

pend_rows = ius_pend + u2u_pend
pend_rows.sort(key=lambda x: (x['route'], x['channel'], x['month']))

all_dump = week_rows + pend_rows

dr = 5
for row_d in all_dump:
    bkt   = row_d['bucket']
    bfill = BUCKET_FILL.get(bkt, PatternFill())
    bclr  = BUCKET_COLOR.get(bkt, '212121')
    vals  = [row_d['route'], bkt, row_d['channel'], row_d['label'],
             row_d['month'], row_d['raise_dt'], row_d['close_dt'],
             row_d['exp'] or None, row_d['act'] or None,
             row_d['sub'], row_d['remarks'], row_d['activity']]
    fmts  = [None, None, None, None, None, None, None, USD_FMT, USD_FMT, None, None, None]
    aligns= [LEFT, LEFT, LEFT, LEFT, LEFT, CENTER, CENTER, RIGHT, RIGHT, LEFT, LEFT, CENTER]

    for col, (v, fm, al) in enumerate(zip(vals, fmts, aligns), 1):
        c = ws_dump.cell(row=dr, column=col, value=v)
        fnt = Font(name='Calibri', size=9,
                   bold=(col == 2),
                   color=(bclr if col == 2 else '212121'))
        style(c, fill=(bfill if col == 2 else None), font=fnt, align=al, fmt=fm, bdr_=bdr())
    ws_dump.row_dimensions[dr].height = 16
    dr += 1

if dr == 5:   # no rows written
    ws_dump.merge_cells('A5:L5')
    ws_dump['A5'] = 'No case activity (raise or close) found in this week range'
    style(ws_dump['A5'], font=Font(name='Calibri', italic=True, color='9E9E9E', size=10), align=CENTER)

dump_widths = [14, 20, 14, 22, 12, 14, 14, 14, 12, 24, 24, 22]
for i, w in enumerate(dump_widths, 1):
    ws_dump.column_dimensions[get_column_letter(i)].width = w
ws_dump.freeze_panes = 'A5'

# ── Save ──────────────────────────────────────────────────────────────────────
os.makedirs(FOLDER, exist_ok=True)
base_out = os.path.join(FOLDER, f'WeeklyRecoveryPivot_W{WEEK_NUM}_{START.strftime("%Y%m%d")}_{END.strftime("%Y%m%d")}.xlsx')
out = base_out
try:
    wb.save(out)
except PermissionError:
    ts  = datetime.now().strftime('%H%M%S')
    out = base_out.replace('.xlsx', f'_{ts}.xlsx')
    wb.save(out)
    print(f"  (file was open in Excel — saved as new copy)")
print(f"\n✅  Saved → {out}")

# ── Auto-save this week's snapshot for next run's Ref column ─────────────────
_all_snaps[END.isoformat()] = {
    'ius': {k: round(v, 2) for k, v in ius_snap_curr.items()},
    'u2u': {k: round(v, 2) for k, v in u2u_snap_curr.items()},
}
with open(SNAP_FILE, 'w') as _f:
    json.dump(_all_snaps, _f, indent=2, sort_keys=True)
print(f"[Snapshot] {END} → saved to snapshots.json (next week's Ref* will auto-load)")

# ── Email alert ───────────────────────────────────────────────────────────────
import base64
from email.mime.multipart import MIMEMultipart
from email.mime.text      import MIMEText
from email.mime.base      import MIMEBase
from email                import encoders
from auth                 import get_gmail_credentials
from googleapiclient.discovery import build as g_build

def send_pivot_alert(excel_path, week_num, start, end,
                     ius_raised_new, ius_closed_old, ius_closed_new,
                     u2u_raised_new, u2u_closed_old, u2u_closed_new):

    gmail_creds = get_gmail_credentials(CF, TF)
    service     = g_build('gmail', 'v1', credentials=gmail_creds)

    # ── Build summary rows for email ────────────────────────────────────────
    def ch_rows_html(route, raised_new, closed_old, closed_new):
        rows = ''
        all_ch = sorted(set(
            list(raised_new.keys()) +
            list(closed_old.keys()) +
            list(closed_new.keys())
        ))
        for ch in all_ch:
            rn     = raised_new.get(ch, empty())
            cn     = closed_new.get(ch, empty())
            by_age = closed_old.get(ch, {})
            co_cnt = sum(len(v['labels']) for v in by_age.values())
            co_amt = sum(v['amt'] for v in by_age.values())
            rows += f'''
            <tr>
              <td style="padding:8px 12px;border-bottom:1px solid #eee;font-weight:600">{route}</td>
              <td style="padding:8px 12px;border-bottom:1px solid #eee">{ch}</td>
              <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center">{len(rn["labels"])}</td>
              <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center">{co_cnt}</td>
              <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:right;color:#2e7d32;font-weight:600">${co_amt:,.2f}</td>
              <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center">{len(cn["labels"])}</td>
            </tr>'''
        return rows

    table_rows = (ch_rows_html('India → US', ius_raised_new, ius_closed_old, ius_closed_new) +
                  ch_rows_html('US → US',    u2u_raised_new, u2u_closed_old, u2u_closed_new))

    html = f"""
    <div style="font-family:Calibri,Arial,sans-serif;max-width:680px;margin:0 auto">
      <div style="background:#1A237E;padding:20px 24px;border-radius:8px 8px 0 0">
        <div style="color:#fff;font-size:18px;font-weight:700">
          📊 Weekly Recovery Pivot — Week {week_num}
        </div>
        <div style="color:rgba(255,255,255,.7);font-size:13px;margin-top:4px">
          {start.strftime('%d %b')} – {end.strftime('%d %b %Y')}
        </div>
      </div>
      <div style="background:#f9f9f9;padding:16px 24px;border:1px solid #e0e0e0;border-top:none">
        <p style="margin:0 0 12px;color:#555;font-size:13px">
          Hey Anand 👋 — your weekly recovery pivot is ready. Find the Excel attached.
        </p>
        <table style="width:100%;border-collapse:collapse;background:#fff;border:1px solid #e0e0e0;border-radius:6px;overflow:hidden">
          <thead>
            <tr style="background:#37474F;color:#fff;font-size:12px">
              <th style="padding:10px 12px;text-align:left">Route</th>
              <th style="padding:10px 12px;text-align:left">Channel</th>
              <th style="padding:10px 12px;text-align:center">Raised<br>This Week</th>
              <th style="padding:10px 12px;text-align:center">Old Cases<br>Closed</th>
              <th style="padding:10px 12px;text-align:right">Backlog<br>Recovered</th>
              <th style="padding:10px 12px;text-align:center">Same-Week<br>Closed</th>
            </tr>
          </thead>
          <tbody>{table_rows}</tbody>
        </table>
        <p style="margin:14px 0 0;font-size:11px;color:#999">
          Auto-generated every Thursday 1 PM IST · Mosaic Wellness Recon Dashboard
        </p>
      </div>
    </div>"""

    msg              = MIMEMultipart('mixed')
    msg['Subject']   = f"📊 Weekly Recovery Pivot Ready — W{week_num} ({start.strftime('%d %b')}–{end.strftime('%d %b %Y')})"
    msg['From']      = CONFIG['email_from']
    msg['To']        = CONFIG['email_from']   # to yourself; add others if needed
    msg.attach(MIMEText(html, 'html'))

    # Attach the Excel file
    with open(excel_path, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition',
                    f'attachment; filename="{os.path.basename(excel_path)}"')
    msg.attach(part)

    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    service.users().messages().send(userId='me', body={'raw': raw}).execute()
    print(f"[Email] Alert sent to {CONFIG['email_from']}")

print(f"[Email] Skipped — file saved to folder only.")
