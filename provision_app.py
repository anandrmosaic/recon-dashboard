import socket
# Force IPv4: on this network, httplib2/googleapiclient tries IPv6 (AAAA) first which times out
_orig_getaddrinfo = socket.getaddrinfo
def _ipv4_first(host, port, family=0, type=0, proto=0, flags=0):
    results = _orig_getaddrinfo(host, port, family, type, proto, flags)
    return sorted(results, key=lambda r: r[0] != socket.AF_INET)
socket.getaddrinfo = _ipv4_first

import json
from flask import Flask, jsonify, render_template, request
from auth import get_credentials
from provision_engine import get_sheet_carriers, process_provision

app = Flask(__name__)

with open('config.json') as f:
    CONFIG = json.load(f)

_creds = None

def get_creds():
    global _creds
    if _creds is None:
        _creds = get_credentials(CONFIG['credentials_file'], CONFIG['token_file'])
    return _creds


@app.route('/')
def provision():
    return render_template('provision.html')


@app.route('/api/provision/detect-carriers', methods=['GET', 'POST'])
def api_detect_carriers():
    try:
        month = int(request.args.get('month') or request.form.get('month', 0))
        year  = int(request.args.get('year')  or request.form.get('year',  0))
        if not month or not year:
            return jsonify({'status': 'ok', 'carriers': []})
        carriers = get_sheet_carriers(get_creds(), CONFIG['sheet_id'], month, year)
        return jsonify({'status': 'ok', 'carriers': carriers})
    except Exception as e:
        import traceback
        print(f"[Provision] detect-carriers error: {traceback.format_exc()}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/provision/generate', methods=['POST'])
def api_provision_generate():
    f = request.files.get('finance_file')
    if not f:
        return jsonify({'status': 'error', 'message': 'No Finance file uploaded'}), 400
    try:
        import openpyxl, io
        finance_bytes = f.read()
        wb = openpyxl.load_workbook(io.BytesIO(finance_bytes), data_only=True)
        print(f"[Provision] Finance file: {f.filename}, sheets: {wb.sheetnames}, cols: {wb.active.max_column}")

        month = int(request.form.get('month', 1))
        year  = int(request.form.get('year', 2026))
        carrier_rates = json.loads(request.form.get('carrier_rates', '{}'))

        result = process_provision(
            finance_bytes, month, year,
            get_creds(), CONFIG['sheet_id'], carrier_rates
        )
        return jsonify({'status': 'ok', **result})
    except Exception as e:
        import traceback
        print(f"[Provision] ERROR: {traceback.format_exc()}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


if __name__ == '__main__':
    print("[Provision] Running at http://localhost:5001")
    print("[Provision] Loading credentials...")
    get_creds()
    print("[Provision] Ready.")
    app.run(debug=False, host='0.0.0.0', port=5001, threaded=True)
