import os
from googleapiclient.discovery import build
from collections import defaultdict
from datetime import datetime, date as date_type, timedelta

MONTH_ORDER = ['January','February','March','April','May','June',
               'July','August','September','October','November','December']


def safe_float(val):
    try:
        return float(str(val).replace(',', '').strip()) if val else 0.0
    except:
        return 0.0


def get_sheet_data(creds, sheet_id, awb_tab, recon_tab=None, data_since=None):
    service = build('sheets', 'v4', credentials=creds)

    # Read AWB tracker raw data
    awb_result = service.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range=f"'{awb_tab}'!A1:AQ8000"
    ).execute()
    awb_values = awb_result.get('values', [])

    # Optionally read remarks from recon pivot tab
    remarks = {}
    if recon_tab:
        try:
            recon_result = service.spreadsheets().values().get(
                spreadsheetId=sheet_id,
                range=f"'{recon_tab}'!A1:R80"
            ).execute()
            recon_values = recon_result.get('values', [])
            remarks_start = next(
                (i for i, r in enumerate(recon_values) if r and 'Channel remark' in str(r[0])), None
            )
            if remarks_start is not None:
                remarks = parse_remarks(recon_values, remarks_start)
        except Exception as e:
            print(f"[Sheets] Could not read remarks: {e}")

    return parse_awb_data(awb_values, remarks, data_since=data_since)


def _parse_date_only(date_str):
    if not date_str or not str(date_str).strip():
        return None
    ds = str(date_str).strip()
    for fmt in ('%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d %b %Y', '%d %B %Y'):
        try:
            return datetime.strptime(ds, fmt).date()
        except ValueError:
            continue
    return None


def _parse_aging(date_str):
    """Return (days_open, bucket_label) from a case raise date string."""
    d = _parse_date_only(date_str)
    if d is None:
        return None, None
    days = (date_type.today() - d).days
    if days < 0:
        return days, None
    if days <= 30:   return days, '0-30 days'
    elif days <= 60: return days, '31-60 days'
    elif days <= 90: return days, '61-90 days'
    else:            return days, '90+ days'


def _parse_resolution(raise_date_str, close_date_str, status):
    """Return (days_to_close, is_closed, is_rejected).
    Uses close_date - raise_date when both dates present.
    Falls back to status keywords for is_closed when close date is missing.
    is_rejected = True when status contains 'reject' (closed but claim denied).
    'No Discrepancy' statuses are excluded — close date present but no real resolution."""
    status_lower = status.lower() if status else ''
    is_rejected = 'reject' in status_lower
    # No discrepancy = case closed because no issue was found, not a real win or loss
    no_discrepancy = 'no' in status_lower and 'discrep' in status_lower
    is_closed_status = is_rejected or bool(any(
        k in status_lower for k in ['close', 'done', 'reimb', 'resolved', 'completed']
    ))
    raise_d = _parse_date_only(raise_date_str)
    close_d = _parse_date_only(close_date_str)
    if raise_d and close_d:
        return max(0, (close_d - raise_d).days), not no_discrepancy, is_rejected
    return None, is_closed_status and not no_discrepancy, is_rejected


def parse_awb_data(values, remarks=None, data_since=None):
    # Cutoff: skip rows before (since_year, since_month)
    since_year  = int(data_since['year'])  if data_since else None
    since_month = int(data_since['month']) if data_since else None
    # Find the header row (col 0 = "Month", col 1 = "Year")
    header_idx = None
    for i, row in enumerate(values):
        if row and str(row[0]).strip().lower() == 'month' and len(row) > 1 and str(row[1]).strip().lower() == 'year':
            header_idx = i
            break

    if header_idx is None:
        print("[Sheets] Header row not found in AWB tracker")
        return {'channel_data': {}, 'transporter_data': {}, 'remarks': remarks or {}, 'kpis': {}}

    # Detect key columns dynamically from header row (exact / keyword matching)
    header_row = values[header_idx]
    headers_lower = [str(c).strip().lower() for c in header_row]

    def col_exact(name, fallback):
        """Find column whose header is exactly `name` (case-insensitive)."""
        try:
            return headers_lower.index(name.lower())
        except ValueError:
            return fallback

    def col_contains(keyword, fallback):
        """Find first column whose header contains `keyword` (case-insensitive)."""
        for i, h in enumerate(headers_lower):
            if keyword in h:
                return i
        return fallback

    # ── All columns detected by header name — immune to column insertions ──
    channel_col      = col_exact('channel',                    fallback=36)
    case_raise_col   = col_contains('case raise',              fallback=None)
    case_close_col   = col_contains('case close',              fallback=None)
    ship_status_col      = col_contains('ship partner portal',  fallback=None)
    if ship_status_col is None:
        ship_status_col  = col_contains('current status',      fallback=None)
    actual_delivery_col  = col_contains('actual delivery',     fallback=None)
    pickup_date_col      = col_contains('pick up',             fallback=None)
    awb_col          = col_contains('shipment awb',            fallback=5)
    platform_col     = col_contains('platform label',          fallback=6)
    transporter_col  = col_exact('transporter',                fallback=999)
    product_col      = col_exact('product name',               fallback=12)
    uniware_col      = col_contains('uniware',                 fallback=13)
    invoice_no_col   = col_contains('invoice no',              fallback=15)
    qty_sent_col     = col_contains('qty sent',                fallback=16)
    lost_stock_col   = col_contains('lost stock',              fallback=28)
    expected_col     = col_contains('expected reimbursement',  fallback=29)
    actual_col       = col_contains('actual reimbursement',    fallback=30)
    reimb_status_col = col_exact('reimbursement status',       fallback=31)
    remark_col       = col_exact('remark',                     fallback=33)

    print(f"[Sheets] Cols — ch:{channel_col} awb:{awb_col} qty:{qty_sent_col} lost:{lost_stock_col} "
          f"exp:{expected_col} act:{actual_col} status:{reimb_status_col} remark:{remark_col} "
          f"raise:{case_raise_col} close:{case_close_col}")

    EXCLUDE_STATUSES = {'abandon', 'rto'}

    print(f"[Sheets] Columns — channel:{channel_col}  case_raise:{case_raise_col}  case_close:{case_close_col}  ship_status:{ship_status_col}  pickup_date:{pickup_date_col}")

    # Weekly buckets: this_week = last 7 days, last_week = 7-14 days ago (by pickup date)
    today_d    = date_type.today()
    week_start = today_d - timedelta(days=7)
    prev_start = today_d - timedelta(days=14)
    weekly = {
        'this_week': {'shipments': 0, 'lost_stock': 0, 'expected': 0.0, 'actual': 0.0},
        'last_week': {'shipments': 0, 'lost_stock': 0, 'expected': 0.0, 'actual': 0.0},
    }

    # Aggregate raw rows
    # Key: (month_str, year_int)  →  channel  →  metrics
    ch_agg = defaultdict(lambda: defaultdict(lambda: {
        'qty_sent': 0.0, 'lost_stock': 0.0, 'expected_reimburs': 0.0, 'actual_reimbursed': 0.0, 'shipment_count': 0
    }))
    tr_agg = defaultdict(lambda: defaultdict(lambda: {'qty_sent': 0.0, 'lost_stock': 0.0}))

    period_set = set()  # (month_str, year_int)
    discrepancies = []  # individual shipment rows where lost_stock > 0
    awb_transporter = {}  # full AWB → transporter map for ALL rows

    for row_offset, row in enumerate(values[header_idx + 1:]):
        if not row or not str(row[0]).strip():
            continue
        month = str(row[0]).strip()
        if month.lower() in ['month', 'grand total']:
            continue
        if month not in MONTH_ORDER:
            continue

        year_raw = str(row[1]).strip() if len(row) > 1 else ''
        if not year_raw.isdigit():
            continue
        year = int(year_raw)

        # Skip rows before the configured cutoff (archive filter)
        if since_year is not None:
            month_idx = MONTH_ORDER.index(month) + 1  # 1-based
            if (year < since_year) or (year == since_year and month_idx < since_month):
                continue

        qty_sent        = safe_float(row[qty_sent_col]   if len(row) > qty_sent_col   else 0)
        lost_stock      = safe_float(row[lost_stock_col] if len(row) > lost_stock_col else 0)
        expected        = safe_float(row[expected_col]   if len(row) > expected_col   else 0)
        actual          = safe_float(row[actual_col]     if len(row) > actual_col     else 0)
        channel         = _norm_channel(str(row[channel_col]).strip()) if len(row) > channel_col else ''
        transporter     = str(row[transporter_col]).strip()  if len(row) > transporter_col  else ''
        platform_label  = str(row[platform_col]).strip()     if len(row) > platform_col     else ''
        # Build full AWB→transporter map — normalize newlines/spaces for reliable matching
        _awb_raw = str(row[awb_col]).strip() if len(row) > awb_col else ''
        if _awb_raw and transporter:
            # Store normalized key (newlines → space, lowercase, stripped)
            for _part in _awb_raw.replace('\n', ' ').split():
                if _part:
                    awb_transporter[_part.lower()] = transporter
            # Also store the full normalized string
            awb_transporter[_awb_raw.replace('\n', ' ').lower().strip()] = transporter
        product_name    = str(row[product_col]).strip()      if len(row) > product_col      else ''
        uniware_code    = str(row[uniware_col]).strip()       if len(row) > uniware_col      else ''
        invoice_no      = str(row[invoice_no_col]).strip()   if len(row) > invoice_no_col   else ''

        period_set.add((month, year))

        if channel:
            d = ch_agg[(month, year)][channel]
            d['qty_sent']          += qty_sent
            d['lost_stock']        += lost_stock
            d['expected_reimburs'] += expected
            d['actual_reimbursed'] += actual
            ship_status = (
                str(row[ship_status_col]).strip().lower()
                if ship_status_col is not None and len(row) > ship_status_col
                else ''
            )
            if ship_status not in EXCLUDE_STATUSES:
                d['shipment_count'] += 1

        # Weekly bucketing by pickup date
        if pickup_date_col is not None and len(row) > pickup_date_col:
            pickup_d = _parse_date_only(str(row[pickup_date_col]).strip())
            if pickup_d is not None:
                if week_start <= pickup_d < today_d:
                    bucket = weekly['this_week']
                elif prev_start <= pickup_d < week_start:
                    bucket = weekly['last_week']
                else:
                    bucket = None
                if bucket is not None:
                    if platform_label:
                        bucket['shipments'] += 1
                    bucket['lost_stock'] += lost_stock
                    bucket['expected']   += expected
                    bucket['actual']     += actual

        if transporter:
            t = tr_agg[(month, year)][transporter]
            t['qty_sent']   += qty_sent

        # Extract case raise date before the condition so it can be used as a trigger
        case_raise_raw = (
            str(row[case_raise_col]).strip()
            if case_raise_col is not None and len(row) > case_raise_col
            else ''
        )

        # Include any row where a case was raised (has raise date) or has lost stock —
        # covers Excess Receive / Inventory Relocated rows that have lost_stock=0
        if (lost_stock > 0 or case_raise_raw or actual > 0) and channel:
            case_close_raw = (
                str(row[case_close_col]).strip()
                if case_close_col is not None and len(row) > case_close_col
                else ''
            )
            reimb_status = str(row[reimb_status_col]).strip() if len(row) > reimb_status_col else ''
            days_open, aging_bucket = _parse_aging(case_raise_raw)
            days_to_close, is_closed, is_rejected = _parse_resolution(case_raise_raw, case_close_raw, reimb_status)
            # Split recovery: carrier vs channel
            is_carrier = 'carrier' in reimb_status.lower()
            carrier_recovered = round(actual, 2) if is_carrier else 0.0
            channel_recovered = 0.0 if is_carrier else round(actual, 2)
            discrepancies.append({
                'row_index':            header_idx + 2 + row_offset,
                'month':                f"{month} {year}",
                'awb':                  str(row[awb_col]).strip()      if len(row) > awb_col      else '',
                'platform_label':       str(row[platform_col]).strip() if len(row) > platform_col else '',
                'transporter':          transporter,
                'channel':              channel,
                'qty_sent':             int(qty_sent),
                'lost_stock':           int(lost_stock),
                'expected_reimburs':    round(expected, 2),
                'actual_reimbursed':    round(actual, 2),
                'carrier_recovered':    carrier_recovered,
                'channel_recovered':    channel_recovered,
                'pending':              round(expected - actual, 2),
                'reimbursement_status': reimb_status,
                'remark':               str(row[remark_col]).strip() if len(row) > remark_col else '',
                'product_name':         product_name,
                'uniware_code':         uniware_code,
                'invoice_no':           invoice_no,
                'case_raise_date':      case_raise_raw,
                'case_close_date':      case_close_raw,
                'days_open':            days_open,
                'aging_bucket':         aging_bucket,
                'days_to_close':        days_to_close,
                'is_closed':            is_closed,
                'is_rejected':          is_rejected,
                'actual_delivery_date': str(row[actual_delivery_col]).strip() if actual_delivery_col is not None and len(row) > actual_delivery_col else '',
                'ship_partner_status':  str(row[ship_status_col]).strip()     if ship_status_col     is not None and len(row) > ship_status_col     else '',
            })
            if transporter:
                t['lost_stock'] += lost_stock

    # Sort periods chronologically: year ASC, then calendar month order
    sorted_periods = sorted(period_set, key=lambda x: (x[1], MONTH_ORDER.index(x[0])))
    period_labels  = [f"{m} {y}" for m, y in sorted_periods]  # e.g. "April 2024"

    # Collect all unique channels and transporters
    all_channels     = sorted({ch for period_data in ch_agg.values() for ch in period_data})
    all_transporters = sorted({tr for period_data in tr_agg.values() for tr in period_data if tr})

    # Build per-channel monthly arrays
    channels = {}
    for ch in all_channels:
        channels[ch] = [
            {
                'qty_sent':          ch_agg[p].get(ch, {}).get('qty_sent', 0.0),
                'lost_stock':        ch_agg[p].get(ch, {}).get('lost_stock', 0.0),
                'expected_reimburs': ch_agg[p].get(ch, {}).get('expected_reimburs', 0.0),
                'actual_reimbursed': ch_agg[p].get(ch, {}).get('actual_reimbursed', 0.0),
                'shipment_count':    ch_agg[p].get(ch, {}).get('shipment_count', 0),
            }
            for p in sorted_periods
        ]

    # Grand total per period
    grand_total = []
    for p in sorted_periods:
        row = {'qty_sent': 0.0, 'lost_stock': 0.0, 'expected_reimburs': 0.0, 'actual_reimbursed': 0.0}
        for ch_data in channels.values():
            r = ch_data[sorted_periods.index(p)]
            for k in row:
                row[k] += r[k]
        grand_total.append(row)

    totals = {k: sum(r[k] for r in grand_total)
              for k in ['qty_sent', 'lost_stock', 'expected_reimburs', 'actual_reimbursed']}

    # Build per-transporter monthly arrays
    transporters = {}
    for tr in all_transporters:
        transporters[tr] = [
            {
                'qty_sent':   tr_agg[p].get(tr, {}).get('qty_sent', 0.0),
                'lost_stock': tr_agg[p].get(tr, {}).get('lost_stock', 0.0),
            }
            for p in sorted_periods
        ]

    channel_data = {
        'months':      period_labels,
        'channels':    channels,
        'grand_total': grand_total,
        'totals':      totals,
    }
    transporter_data = {
        'months':       period_labels,
        'transporters': transporters,
        'totals':       {},
    }

    return {
        'channel_data':     channel_data,
        'transporter_data': transporter_data,
        'remarks':          remarks or {},
        'kpis':             calculate_kpis(channel_data),
        'discrepancies':    discrepancies,
        'weekly':           weekly,
        'awb_transporter':  awb_transporter,
    }


def parse_remarks(values, start):
    remarks = {}
    current_channel = current_month = None
    months = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']

    for row in values[start + 1:]:
        if not row or not str(row[0]).strip():
            continue
        cell = str(row[0]).strip()
        if cell.lower() in ['tiktok', 'shipbob', 'amazon']:
            current_channel = _norm_channel(cell)
            remarks.setdefault(current_channel, {})
        elif cell in months:
            current_month = cell
        elif current_channel and current_month and len(cell) > 5:
            remarks[current_channel][current_month] = cell

    return remarks


def calculate_kpis(channel_data):
    t = channel_data.get('totals', {})
    total_shipped = t.get('qty_sent', 0)
    total_lost    = t.get('lost_stock', 0)
    expected      = t.get('expected_reimburs', 0)
    actual        = t.get('actual_reimbursed', 0)
    pending       = expected - actual
    recovery_rate = round((actual / expected * 100), 1) if expected > 0 else 0
    loss_rate     = round((total_lost / total_shipped * 100), 3) if total_shipped > 0 else 0

    return {
        'total_shipped':    int(total_shipped),
        'total_lost':       int(total_lost),
        'expected_recovery': round(expected, 2),
        'actual_recovered': round(actual, 2),
        'pending_recovery': round(pending, 2),
        'recovery_rate':    recovery_rate,
        'loss_rate':        loss_rate,
    }


def get_ups_claims_data(creds, sheet_id):
    """Read UPS Claim tab + AWB Master (UPS) tab and return structured claims data."""
    if not sheet_id:
        return {'summary': {}, 'claims': []}
    service = build('sheets', 'v4', credentials=creds)

    # Read AWB Master (UPS) - col A = AWB, col B = TRUE/FALSE, row 1 has counts in cols F-H
    master_result = service.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range="'AWB Master (ups)'!A1:H1000"
    ).execute()
    master_rows = master_result.get('values', [])

    total_awbs = 0
    claim_filed = 0
    not_filed = 0
    if master_rows:
        header = master_rows[0]
        # Counts are in row 1 cols F(5), G(6), H(7)
        if len(master_rows) > 1 and len(master_rows[1]) >= 8:
            try: claim_filed = int(str(master_rows[1][5]).replace(',','').strip())
            except: pass
            try: not_filed = int(str(master_rows[1][6]).replace(',','').strip())
            except: pass
            try: total_awbs = int(str(master_rows[1][7]).replace(',','').strip())
            except: pass

    # Count AWBs with FALSE tracking (col B = FALSE → no claim form received)
    false_tracking_awbs = []
    for r in master_rows[1:]:
        if len(r) >= 2 and str(r[1]).strip().upper() == 'FALSE' and str(r[0]).strip():
            false_tracking_awbs.append(str(r[0]).strip())
    false_tracking_count = len(false_tracking_awbs)

    # Read UPS Claim tab — col H = remark/notes
    claim_result = service.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range="'UPS Claim'!A1:H500"
    ).execute()
    claim_rows = claim_result.get('values', [])

    claims = []
    if len(claim_rows) > 1:
        for row in claim_rows[1:]:
            if not row or not str(row[0]).strip():
                continue
            padded = row + [''] * max(0, 8 - len(row))
            parent_awb    = str(padded[0]).strip()
            lost_awb      = str(padded[1]).strip()
            lost_qty      = str(padded[2]).strip()
            claim_amount  = str(padded[3]).strip()
            form_received = str(padded[4]).strip()
            approved_date = str(padded[5]).strip()
            settled_date  = str(padded[6]).strip()
            remark        = str(padded[7]).strip()

            remark_lower = remark.lower()
            # Determine state — priority order is critical
            # 1. Declined first (remark is definitive, regardless of claim_amount)
            # 2. Settled = claim_amount AND UTR (SCBLH prefix) present in remark
            # 3. Approved-not-settled = claim_amount BUT no UTR received yet
            # 4. Filed pending = form received, no amount yet
            # 5. Not filed
            if 'claim declined' in remark_lower or \
               ('package deliver' in remark_lower and 'ups' in remark_lower):
                state = 'declined'
            elif claim_amount and 'scblh' in remark_lower:
                state = 'settled'               # UTR in remark = money transferred
            elif claim_amount:
                state = 'approved_not_settled'  # approved but awaiting bank credit
            elif form_received:
                state = 'filed_pending'
            else:
                state = 'not_filed'

            claims.append({
                'parent_awb':    parent_awb,
                'lost_awb':      lost_awb,
                'lost_qty':      lost_qty,
                'claim_amount':  claim_amount,
                'form_received': form_received,
                'approved_date': approved_date,
                'settled_date':  settled_date,
                'remark':        remark,
                'state':         state,
            })

    settled_count              = sum(1 for c in claims if c['state'] == 'settled')
    approved_not_settled_count = sum(1 for c in claims if c['state'] == 'approved_not_settled')
    filed_pending_count        = sum(1 for c in claims if c['state'] == 'filed_pending')
    declined_count             = sum(1 for c in claims if c['state'] == 'declined')

    return {
        'summary': {
            'total_awbs':                  total_awbs,
            'claim_filed':                 claim_filed,
            'not_filed':                   not_filed,
            'false_tracking_count':        false_tracking_count,
            'false_tracking_awbs':         false_tracking_awbs,
            'settled_count':               settled_count,
            'approved_not_settled_count':  approved_not_settled_count,
            'filed_pending_count':         filed_pending_count,
            'declined_count':              declined_count,
        },
        'claims': claims,
    }


def get_recon_recovery_totals(creds, sheet_id, tab_name):
    """Directly sum Expected + Actual Reimbursement columns from recon sheet.
    Bypasses parse_awb_data to get exact totals matching what user sees in sheet."""
    service = build('sheets', 'v4', credentials=creds)
    result  = service.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range=f"'{tab_name}'!A1:AZ5000"
    ).execute()
    values = result.get('values', [])
    if not values:
        return {}

    # Find header row
    header_row = None
    for row in values:
        if row and str(row[0]).strip().lower() == 'month':
            header_row = row
            break
    if not header_row:
        return {}

    headers = [str(h).strip().lower() for h in header_row]

    # Find Expected, Actual Reimbursement and Lost Stock columns
    expected_idx  = next((i for i, h in enumerate(headers) if 'expected reimburs' in h), None)
    actual_idx    = next((i for i, h in enumerate(headers) if 'actual reimburs'   in h), None)
    lost_idx      = next((i for i, h in enumerate(headers) if 'lost stock'        in h), None)

    if expected_idx is None or actual_idx is None:
        return {}

    expected_total = 0.0
    actual_total   = 0.0
    lost_total     = 0.0
    header_found   = False

    for row in values:
        if not header_found:
            if row and str(row[0]).strip().lower() == 'month':
                header_found = True
            continue
        if not row or not str(row[0]).strip():
            continue
        try:
            if len(row) > expected_idx:
                v = str(row[expected_idx]).replace(',', '').strip()
                if v: expected_total += float(v)
        except: pass
        try:
            if len(row) > actual_idx:
                v = str(row[actual_idx]).replace(',', '').strip()
                if v: actual_total += float(v)
        except: pass
        try:
            if lost_idx is not None and len(row) > lost_idx:
                v = str(row[lost_idx]).replace(',', '').strip()
                if v: lost_total += float(v)
        except: pass

    return {
        'expected_reimburs':  round(expected_total, 2),
        'actual_reimbursed':  round(actual_total,   2),
        'lost_stock':         round(lost_total,     2),
    }


def get_us2us_data(creds, sheet_id):
    """Read Internal US 2 US tab — classify rows by Bucket + SubBucket."""
    service = build('sheets', 'v4', credentials=creds)
    result  = service.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range="'Internal US 2 US'!A1:AZ1000"
    ).execute()
    values = result.get('values', [])
    if len(values) < 4:
        return {'kpis': {}, 'rows': [], 'monthly': {}, 'months': []}

    # Row 3 (index 2) = headers; data starts row 4 (index 3)
    raw_hdrs = values[2]
    headers  = [str(h).strip().lower() for h in raw_hdrs]

    def fc(kw):
        for i, h in enumerate(headers):
            if kw in h: return i
        return -1

    year_col     = fc('year')
    month_col    = fc('month')
    from_ch_col  = fc('from channel')
    to_ch_col    = fc('to channel')
    from_lbl_col = fc('from (label')
    to_lbl_col   = fc('to (label')
    qty_col       = fc('quantity')
    carrier_col   = fc('carrier')
    status_col    = fc('channel grn status')
    reimb_col     = fc('reimbursement status')
    bucket_col    = fc('bucket')
    subbucket_col = fc('subbucket')
    diff_col      = fc('sum difference')
    exp_col       = fc('expected reimburs')
    act_col       = fc('amt recov')
    raise_col     = fc('case raised date')
    resolve_col   = fc('case resolve date')
    freight_col   = fc('freight')
    grn_col       = fc('inwarded by channel')
    sku_col       = fc('sku')

    print(f"[US2US] Cols — month:{month_col} bucket:{bucket_col} sub:{subbucket_col} "
          f"exp:{exp_col} act:{act_col} from:{from_ch_col} to:{to_ch_col}")

    kpis = {
        'closed':       {'count': 0, 'qty': 0, 'expected': 0.0, 'recovered': 0.0, 'sub': {}},
        'in_progress':  {'count': 0, 'qty': 0, 'expected': 0.0, 'recovered': 0.0},
        'pending':      {'count': 0, 'qty': 0, 'expected': 0.0},
        'intransit':    {'count': 0, 'qty': 0},
    }
    # Unique From-Label sets per bucket (col G) — same label can span multiple SKU rows
    _seen_us = {
        'closed': set(), 'in_progress': set(), 'pending': set(), 'intransit': set(),
    }
    _monthly_seen_us = defaultdict(lambda: defaultdict(set))  # month -> bucket_key -> labels
    monthly = defaultdict(lambda: {
        'closed': 0, 'in_progress': 0, 'pending': 0, 'intransit': 0,
        'expected': 0.0, 'recovered': 0.0, 'qty': 0
    })
    rows = []

    def g(row, idx):
        return row[idx] if idx >= 0 and len(row) > idx else ''

    for i, row in enumerate(values[3:]):
        if not row: continue
        month = str(g(row, month_col)).strip()
        if month not in MONTH_ORDER: continue

        bucket    = str(g(row, bucket_col)).strip()
        subbucket = str(g(row, subbucket_col)).strip()
        bl        = bucket.lower()
        sl        = subbucket.lower()
        exp       = safe_float(g(row, exp_col))
        act       = safe_float(g(row, act_col))
        qty       = safe_float(g(row, qty_col))
        diff      = safe_float(g(row, diff_col))

        from_label = str(g(row, from_lbl_col)).strip()

        # ── Classify ──────────────────────────────────────────────────────────
        if 'closed' in bl or 'resolved' in bl or 'received' in bl:
            _seen_us['closed'].add(from_label)
            _monthly_seen_us[month]['closed'].add(from_label)
            kpis['closed']['qty']       += qty
            kpis['closed']['expected']  += exp
            kpis['closed']['recovered'] += act
            kpis['closed']['sub'][subbucket] = kpis['closed']['sub'].get(subbucket, 0) + 1
            bucket_key = 'closed'
        elif 'transit' in bl or 'transit' in sl:
            _seen_us['intransit'].add(from_label)
            _monthly_seen_us[month]['intransit'].add(from_label)
            kpis['intransit']['qty']   += qty
            bucket_key = 'intransit'
        elif 'progress' in bl or 'progress' in sl or 'raised' in sl:
            _seen_us['in_progress'].add(from_label)
            _monthly_seen_us[month]['in_progress'].add(from_label)
            kpis['in_progress']['qty']       += qty
            kpis['in_progress']['expected']  += exp
            kpis['in_progress']['recovered'] += act
            # Partial reimbursement: claim in progress but some $ already received
            if act > 0 and exp > 0 and act < exp:
                bucket_key = 'partial'   # ← partial: claim raised, balance still pending
            elif act >= exp and exp > 0:
                bucket_key = 'closed'    # ← fully recovered despite label; treat as closed
            else:
                bucket_key = 'in_progress'
        elif diff != 0 or exp > 0:
            _seen_us['pending'].add(from_label)
            _monthly_seen_us[month]['pending'].add(from_label)
            kpis['pending']['qty']      += qty
            kpis['pending']['expected'] += exp
            bucket_key = 'pending'
        else:
            bucket_key = 'other'

        monthly[month]['expected']  += exp
        monthly[month]['recovered'] += act
        monthly[month]['qty']       += qty

        reimb_status   = str(g(row, reimb_col)).strip() if reimb_col >= 0 and len(row) > reimb_col else ''
        is_carrier_pay = 'carrier' in reimb_status.lower()
        rows.append({
            'row_index':            i + 4,
            'month':                month,
            'from_channel':         str(g(row, from_ch_col)).strip(),
            'to_channel':           str(g(row, to_ch_col)).strip(),
            'from_label':           from_label,
            'to_label':             str(g(row, to_lbl_col)).strip(),
            'sku':                  str(g(row, sku_col)).strip()[:40],
            'qty':                  int(qty),
            'grn':                  int(safe_float(g(row, grn_col))),
            'diff':                 int(diff),
            'carrier':              str(g(row, carrier_col)).strip(),
            'status':               str(g(row, status_col)).strip(),
            'reimbursement_status': reimb_status,
            'carrier_recovered':    round(act, 2) if is_carrier_pay else 0.0,
            'channel_recovered':    0.0 if is_carrier_pay else round(act, 2),
            'bucket':               bucket,
            'subbucket':            subbucket,
            'bucket_key':           bucket_key,
            'expected':             round(exp, 2),
            'recovered':            round(act, 2),
            'pending':              round(max(0, exp - act), 2),
            'freight':              round(safe_float(g(row, freight_col)), 2),
            'case_raise_date':      str(g(row, raise_col)).strip(),
            'case_resolve_date':    str(g(row, resolve_col)).strip(),
        })

    months_present = sorted(
        {r['month'] for r in rows if r['month'] in MONTH_ORDER},
        key=lambda x: MONTH_ORDER.index(x)
    )

    # Set unique From-Label counts per bucket (deduplicates same label across multiple SKU rows)
    for bk in ['closed', 'intransit', 'in_progress', 'pending']:
        kpis[bk]['count'] = len(_seen_us[bk])

    # Back-fill monthly counts with unique label counts
    for m, bk_map in _monthly_seen_us.items():
        for bk, labels in bk_map.items():
            monthly[m][bk] = len(labels)

    for k in ['closed', 'in_progress', 'pending']:
        for f in ['expected', 'recovered']:
            if f in kpis[k]:
                kpis[k][f] = round(kpis[k][f], 2)

    print(f"[US2US] Closed:{kpis['closed']['count']} Intransit:{kpis['intransit']['count']} "
          f"Pending:{kpis['pending']['count']} InProgress:{kpis['in_progress']['count']}")

    return {
        'kpis':    kpis,
        'rows':    rows,
        'monthly': {m: dict(v) for m, v in monthly.items()},
        'months':  months_present,
    }


def _norm_channel(ch):
    """Normalise raw channel strings → TikTok | Amazon | ShipBob (or original)."""
    lc = ch.lower().strip()
    if 'tiktok' in lc or 'tik tok' in lc or 'tik-tok' in lc:
        return 'TikTok'
    if 'amazon' in lc:
        return 'Amazon'
    if 'shipbob' in lc:
        return 'ShipBob'
    return ch  # keep as-is for any future channels


def get_india_us_data(creds, sheet_id):
    """Read Inward India to US tab — classify rows by Bucket + Sub Remarks."""
    service = build('sheets', 'v4', credentials=creds)
    result  = service.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range="'Inward India to US'!A1:AZ2000"
    ).execute()
    values = result.get('values', [])
    if len(values) < 3:
        return {'kpis': {}, 'rows': [], 'monthly': {}, 'months': []}

    # Row 2 (index 1) = headers; data starts row 3 (index 2)
    raw_hdrs = values[1]
    headers  = [str(h).strip().lower() for h in raw_hdrs]

    def fc(kw):
        for i, h in enumerate(headers):
            if kw in h: return i
        return -1

    month_col       = fc('month')
    bucket_col      = fc('bucket')
    sub_col         = fc('sub remark')
    awb_col         = fc('shipment awb')
    label_col       = fc('platform label')
    old_ibr_col     = fc('old ibr')           # parent IBR link for split shipments
    child_col       = fc('child product')
    qty_col         = fc('final sent qty')
    grn_col         = fc('final grn')
    diff_col        = fc('final diff')
    exp_col         = fc('expected reimburs')
    act_col         = fc('actual reimburs')
    status_col      = fc('reimbursement status')
    raise_col       = fc('case raise date')
    close_col       = fc('case close date')
    lost_col        = fc('lost stock')
    # Exact-match 'channel' first — avoids picking up 'Sub Channel', 'Sales Channel', etc.
    # that appear before column AM.  Fallback to substring, then hardcode AM (col 39 = idx 38).
    channel_col     = next((i for i, h in enumerate(headers) if h == 'channel'), -1)
    if channel_col < 0:
        channel_col = fc('channel')
    if channel_col < 0:
        channel_col = 38  # hardcoded: column AM (39th column, 0-indexed 38)
    transporter_col = fc('transporter')
    # Exact-match 'remarks' to avoid matching 'sub remark'; fallback to col AE (idx 30)
    remarks_col     = next((i for i, h in enumerate(headers) if h == 'remarks'), -1)
    if remarks_col < 0:
        remarks_col = next((i for i, h in enumerate(headers) if h == 'remark'), -1)
    if remarks_col < 0:
        remarks_col = 30  # hardcoded: column AE (31st column, 0-indexed 30)

    print(f"[IndiaUS] Cols — month:{month_col} bucket:{bucket_col} sub:{sub_col} "
          f"exp:{exp_col} act:{act_col} qty:{qty_col} channel:{channel_col} remarks:{remarks_col}")

    kpis = {
        'closed':              {'count': 0, 'qty': 0, 'expected': 0.0, 'actual': 0.0, 'sub': {}},
        'intransit':           {'count': 0, 'qty': 0},
        'short_grn_pending':   {'count': 0, 'qty': 0, 'expected': 0.0},
        'short_grn_progress':  {'count': 0, 'qty': 0, 'expected': 0.0, 'actual': 0.0},
    }
    # Unique label sets per bucket (col G = Platform Label) — combos share same label across rows
    _seen = {
        'closed': set(), 'intransit': set(),
        'short_grn_pending': set(), 'short_grn_progress': set(),
    }
    # Per-month unique label sets for monthly count columns
    _monthly_seen = defaultdict(lambda: defaultdict(set))  # month -> bucket_key -> labels
    monthly = defaultdict(lambda: {
        'closed': 0, 'intransit': 0, 'short_grn_pending': 0, 'short_grn_progress': 0,
        'expected': 0.0, 'actual': 0.0, 'qty': 0
    })
    rows = []

    def g(row, idx):
        return row[idx] if idx >= 0 and len(row) > idx else ''

    for i, row in enumerate(values[2:]):
        if not row: continue
        month = str(g(row, month_col)).strip()
        if month not in MONTH_ORDER: continue

        bucket  = str(g(row, bucket_col)).strip()
        sub     = str(g(row, sub_col)).strip()
        exp     = safe_float(g(row, exp_col))
        act     = safe_float(g(row, act_col))
        qty     = safe_float(g(row, qty_col))
        bl      = bucket.lower()
        sl      = sub.lower()

        label = str(g(row, label_col)).strip()

        # ── Classify ──────────────────────────────────────────────────────────
        if bl == 'closed':
            _seen['closed'].add(label)
            _monthly_seen[month]['closed'].add(label)
            kpis['closed']['qty']      += qty
            kpis['closed']['expected'] += exp
            kpis['closed']['actual']   += act
            kpis['closed']['sub'][sub]  = kpis['closed']['sub'].get(sub, 0) + 1
            bucket_key = 'closed'

        elif bl == 'intransit':
            _seen['intransit'].add(label)
            _monthly_seen[month]['intransit'].add(label)
            kpis['intransit']['qty']   += qty
            bucket_key = 'intransit'

        elif bl == 'short grn':
            if 'claim raised but not received' in sl:
                _seen['short_grn_progress'].add(label)
                _monthly_seen[month]['short_grn_progress'].add(label)
                kpis['short_grn_progress']['qty']      += qty
                kpis['short_grn_progress']['expected'] += exp
                kpis['short_grn_progress']['actual']   += act
                bucket_key = 'short_grn_progress'
            elif 'pending qty inwarded against new ibr' in sl:
                # Waiting for child IBRs to cover the shortfall — distinct from plain pending
                _seen['short_grn_pending'].add(label)
                _monthly_seen[month]['short_grn_pending'].add(label)
                kpis['short_grn_pending']['qty']      += qty
                kpis['short_grn_pending']['expected'] += exp
                bucket_key = 'short_grn_awaiting_ibr'  # tagged separately for UI
            else:  # Pending to claim
                _seen['short_grn_pending'].add(label)
                _monthly_seen[month]['short_grn_pending'].add(label)
                kpis['short_grn_pending']['qty']      += qty
                kpis['short_grn_pending']['expected'] += exp
                bucket_key = 'short_grn_pending'
        else:
            bucket_key = 'other'

        monthly[month]['expected'] += exp
        monthly[month]['actual']   += act
        monthly[month]['qty']      += qty

        old_ibr = str(g(row, old_ibr_col)).strip()
        reimb_status   = str(g(row, status_col)).strip() if status_col >= 0 and len(row) > status_col else ''
        is_carrier_pay = 'carrier' in reimb_status.lower()
        rows.append({
            'row_index':            i + 3,
            'month':                month,
            'awb':                  str(g(row, awb_col)).strip(),
            'label':                label,
            'old_ibr':              old_ibr,                           # parent IBR (child rows only)
            'product':              str(g(row, child_col)).strip()[:50],
            'qty':                  int(qty),
            'grn':                  int(safe_float(g(row, grn_col))),
            'diff':                 int(safe_float(g(row, diff_col))),
            'channel':              _norm_channel(str(g(row, channel_col)).strip()),
            'transporter':          str(g(row, transporter_col)).strip(),
            'remarks':              str(g(row, remarks_col)).strip(),
            'reimbursement_status': reimb_status,
            'carrier_recovered':    round(act, 2) if is_carrier_pay else 0.0,
            'channel_recovered':    0.0 if is_carrier_pay else round(act, 2),
            'case_raise_date':      str(g(row, raise_col)).strip(),
            'case_close_date':      str(g(row, close_col)).strip(),
            'lost_stock':           int(safe_float(g(row, lost_col))),
            'expected':             round(exp, 2),
            'actual':               round(act, 2),
            'pending':              round(max(0, exp - act), 2),
            'bucket':               bucket,
            'sub_remark':           sub,
            'bucket_key':           bucket_key,
        })

    months_present = sorted(
        {r['month'] for r in rows if r['month'] in MONTH_ORDER},
        key=lambda x: MONTH_ORDER.index(x)
    )

    # Set unique shipment counts from label sets (combos share same IBR across rows)
    for bk in ['closed', 'intransit', 'short_grn_pending', 'short_grn_progress']:
        kpis[bk]['count'] = len(_seen[bk])

    # Back-fill monthly counts with unique label counts (not row counts)
    for m, bk_map in _monthly_seen.items():
        for bk, labels in bk_map.items():
            monthly[m][bk] = len(labels)

    # Round kpi floats
    for k in ['closed', 'short_grn_pending', 'short_grn_progress']:
        for f in ['expected', 'actual']:
            if f in kpis[k]:
                kpis[k][f] = round(kpis[k][f], 2)

    print(f"[IndiaUS] Closed:{kpis['closed']['count']} Intransit:{kpis['intransit']['count']} "
          f"Pending:{kpis['short_grn_pending']['count']} InProgress:{kpis['short_grn_progress']['count']}")

    return {
        'kpis':    kpis,
        'rows':    rows,
        'monthly': {m: dict(v) for m, v in monthly.items()},
        'months':  months_present,
    }


def get_outward_loss_data(creds, sheet_id):
    if not sheet_id:
        return {'headers': [], 'rows': []}
    from googleapiclient.discovery import build
    service = build('sheets', 'v4', credentials=creds)
    result = service.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range="Sheet1!A1:F5000"
    ).execute()
    values = result.get('values', [])
    if not values:
        return {'headers': [], 'rows': []}
    return {'headers': values[0], 'rows': values[1:]}


# ─────────────────────────────────────────────────────────────────────────────
# ShipBob D2C Claims tab
# Reads by column HEADER NAME — paste any set of columns in any order.
# Recognised headers (case-insensitive, partial match):
#   month, shipment id, sales channel / channel / ingestion channel,
#   sku, sub bucket, claim status, amt, expected claim amount / expected,
#   claims remark / remark, carrier, main bucket, row status
# ─────────────────────────────────────────────────────────────────────────────
def get_shipbob_d2c_data(creds, sheet_id, tab_name='ShipBob D2C Claims'):
    """Read ShipBob D2C claim rows from a Google Sheet tab.

    Column order does not matter — columns are matched by header name so the
    user can paste the full Dump sheet (or a subset) without reordering.
    Only rows where Main Bucket = 'Under Dispute' (or Sub Bucket is a known
    claim bucket) are aggregated; purely Delivered/Intransit rows are ignored.
    """
    service = build('sheets', 'v4', credentials=creds)

    try:
        result = service.spreadsheets().values().get(
            spreadsheetId=sheet_id,
            range=f"'{tab_name}'!A1:BZ200000"
        ).execute()
    except Exception as e:
        print(f"[ShipBob D2C] Sheet read failed ({tab_name}): {e}")
        return _empty_d2c()

    values = result.get('values', [])
    if len(values) < 2:
        return _empty_d2c()

    # ── Build header → column-index map (case-insensitive) ───────────────────
    raw_hdrs = [str(h).strip().lower() if h else '' for h in values[0]]

    def _find(candidates):
        """Return index of first header that contains any candidate string."""
        for h in candidates:
            for i, rh in enumerate(raw_hdrs):
                if h in rh:
                    return i
        return None

    ci_month      = _find(['month'])
    ci_ship_id    = _find(['shipment id'])
    ci_channel    = _find(['sales channel', 'ingestion channel store', 'channel'])
    ci_sku        = _find(['sku'])
    ci_line_name  = _find(['line item name'])
    ci_line_qty   = _find(['line item qty'])
    ci_sub_bucket = _find(['sub bucket'])
    ci_main_bucket= _find(['main bucket'])
    ci_claim_st   = _find(['claim status'])
    ci_amt        = _find(['amt'])
    ci_expected   = _find(['expected claim', 'expected'])
    ci_remark     = _find(['claims remark', 'remark'])
    ci_carrier    = _find(['carrier'])
    ci_row_status = _find(['row status'])
    ci_days       = _find(['days of order'])

    print(f"[ShipBob D2C] Column map: month={ci_month} ship={ci_ship_id} ch={ci_channel} "
          f"sku={ci_sku} sub={ci_sub_bucket} main={ci_main_bucket} "
          f"amt={ci_amt} exp={ci_expected} remark={ci_remark} carrier={ci_carrier}")

    def gv(row, idx):
        if idx is None or idx >= len(row):
            return ''
        v = row[idx]
        return str(v).strip() if v is not None else ''

    # Normalise bucket label → key
    BUCKET_MAP = {
        'claim raised and received':     'rec',
        'claim raised but not received': 'prog',
        'pending to claim':              'pend',
        'claim window expired':          'expired',
        'rto':                           'rto',
        'cancelled':                     'cancelled',
        'delivered':                     'delivered',
        'intransit':                     'intransit',
    }
    SKIP_BUCKETS = {'delivered', 'intransit', 'other'}

    rows          = []
    monthly       = {}   # month → {rec, prog, pend, expired}  (claim rows only)
    ch_agg        = {}   # channel → {rec, prog, pend}          (claim rows only)
    kpis          = {
        'rec':     {'count': 0, 'amt': 0.0},
        'prog':    {'count': 0, 'exp': 0.0},
        'pend':    {'count': 0, 'exp': 0.0},
        'expired': {'count': 0, 'exp': 0.0},
    }
    prog_shipments = set()   # unique shipment IDs for pending-receipt
    pend_shipments = set()   # unique shipment IDs for pending-to-claim
    # Pivot 1 — ALL rows: sub-bucket label × month → shipment count
    pivot1        = {}   # {sub_bucket_label: {month: count}}
    pivot1_months = set()
    # Total qty per shipment (sum across all line items for same Shipment ID)
    shipment_qty  = {}   # {shipment_id: total_qty}

    for raw in values[1:]:
        if not any(raw):
            continue

        month        = gv(raw, ci_month)
        shipment_id  = gv(raw, ci_ship_id)
        channel      = gv(raw, ci_channel)
        sku          = gv(raw, ci_sku)
        line_name    = gv(raw, ci_line_name) or sku
        sub_bucket   = gv(raw, ci_sub_bucket)
        main_bucket  = gv(raw, ci_main_bucket)
        claim_status = gv(raw, ci_claim_st)
        amt          = safe_float(gv(raw, ci_amt))
        exp          = safe_float(gv(raw, ci_expected))
        remark       = gv(raw, ci_remark)
        carrier      = gv(raw, ci_carrier)
        days         = safe_float(gv(raw, ci_days))

        # Determine bucket key — sub_bucket is authoritative, fall back to main
        bk = BUCKET_MAP.get(sub_bucket.strip().lower(),
             BUCKET_MAP.get(main_bucket.strip().lower(), 'other'))

        # ── Pivot 1: count ALL rows by sub-bucket label × month ──────────
        sb_label = (sub_bucket or main_bucket or 'Unknown').strip()
        if month and sb_label:
            if sb_label not in pivot1:
                pivot1[sb_label] = {}
            pivot1[sb_label][month] = pivot1[sb_label].get(month, 0) + 1
            pivot1_months.add(month)

        # ── Shipment qty: sum line item qty per Shipment ID ───────────────
        if shipment_id:
            qty_val = safe_float(gv(raw, ci_line_qty))
            shipment_qty[shipment_id] = shipment_qty.get(shipment_id, 0) + qty_val

        # Skip rows that are purely operational (no claim context)
        if bk in SKIP_BUCKETS:
            continue
        # Also skip if there's no claim amount AND not a known claim bucket
        claim_bks = {'rec', 'prog', 'pend', 'expired'}
        if bk not in claim_bks and amt == 0 and exp == 0:
            continue

        row = {
            'month':        month,
            'shipment_id':  shipment_id,
            'channel':      channel,
            'sku':          sku,
            'line_name':    line_name,
            'sub_bucket':   sub_bucket or main_bucket,
            'bucket_key':   bk,
            'claim_status': claim_status,
            'amt':          round(amt, 2),
            'exp':          round(exp, 2),
            'remark':       remark,
            'carrier':      carrier,
            'days':         int(days) if days else 0,
            # total_qty filled after full-pass completes
        }
        rows.append(row)

        # ── Monthly aggregation ──
        if month:
            if month not in monthly:
                monthly[month] = {'rec': 0.0, 'prog': 0.0, 'pend': 0.0, 'expired': 0.0}
            if bk == 'rec':
                monthly[month]['rec']     += amt
            elif bk == 'prog':
                monthly[month]['prog']    += exp
            elif bk == 'pend':
                monthly[month]['pend']    += exp
            elif bk == 'expired':
                monthly[month]['expired'] += exp

        # ── KPI aggregation ──
        if bk in kpis:
            if bk == 'rec':
                kpis[bk]['count'] += 1
                kpis[bk]['amt'] += amt
            elif bk == 'prog':
                prog_shipments.add(shipment_id)
                kpis[bk]['exp'] += exp
            elif bk == 'pend':
                pend_shipments.add(shipment_id)
                kpis[bk]['exp'] += exp
            elif bk == 'expired':
                kpis[bk]['count'] += 1
                kpis[bk]['exp'] += exp

        # ── Channel aggregation ──
        if channel:
            if channel not in ch_agg:
                ch_agg[channel] = {'rec': 0.0, 'prog': 0.0, 'pend': 0.0}
            if bk == 'rec':
                ch_agg[channel]['rec'] += amt
            elif bk == 'prog':
                ch_agg[channel]['prog'] += exp
            elif bk == 'pend':
                ch_agg[channel]['pend'] += exp

    # Unique order counts for prog / pend
    kpis['prog']['count'] = len(prog_shipments)
    kpis['pend']['count'] = len(pend_shipments)

    # Backfill total_qty into every claim row
    for row in rows:
        row['total_qty'] = int(shipment_qty.get(row['shipment_id'], 0))

    months = sorted(monthly.keys(),
                    key=lambda m: MONTH_ORDER.index(m) if m in MONTH_ORDER else 99)
    p1_months = sorted(pivot1_months,
                       key=lambda m: MONTH_ORDER.index(m) if m in MONTH_ORDER else 99)

    # Round all monthly values
    for m in monthly:
        for k in monthly[m]:
            monthly[m][k] = round(monthly[m][k], 2)
    for ch in ch_agg:
        for k in ch_agg[ch]:
            ch_agg[ch][k] = round(ch_agg[ch][k], 2)
    for bk in kpis:
        for k in kpis[bk]:
            if isinstance(kpis[bk][k], float):
                kpis[bk][k] = round(kpis[bk][k], 2)

    return {
        'rows':         rows,
        'monthly':      monthly,
        'months':       months,
        'kpis':         kpis,
        'channels':     ch_agg,
        'pivot1':       pivot1,
        'pivot1_months': p1_months,
    }


def _empty_d2c():
    return {'rows': [], 'monthly': {}, 'months': [], 'kpis': {}, 'channels': {},
            'pivot1': {}, 'pivot1_months': []}


# ─────────────────────────────────────────────────────────────────────────────
# ShipBob D2C Claims — read from local Excel file (G: Drive folder)
# Automatically picks the latest "Shipbob Order Data - Updated *.xlsx" file.
# Falls back to Sheets tab if folder is missing or no file found.
# ─────────────────────────────────────────────────────────────────────────────
def get_shipbob_d2c_from_excel(folder):
    """Read ShipBob D2C claim data from the latest Excel file in *folder*.

    Returns the same dict structure as get_shipbob_d2c_data() so the rest of
    the app needs no changes.
    """
    import glob as _glob
    import openpyxl as _openpyxl

    pattern = os.path.join(folder, "Shipbob Order Data - Updated *.xlsx")
    files   = sorted(_glob.glob(pattern), key=os.path.getmtime)
    if not files:
        print(f"[ShipBob Excel] No matching file in {folder!r} — falling back")
        return None   # caller will fall back to Sheets tab

    latest = files[-1]
    print(f"[ShipBob Excel] Reading {os.path.basename(latest)}")

    try:
        wb = _openpyxl.load_workbook(latest, read_only=True, data_only=True)
    except Exception as e:
        print(f"[ShipBob Excel] Failed to open workbook: {e}")
        return None

    # Try 'Dump' sheet first, then active sheet
    ws = wb["Dump"] if "Dump" in wb.sheetnames else wb.active

    raw_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if len(raw_rows) < 2:
        return None

    # ── Header map (same partial-match logic as Sheets version) ───────────────
    raw_hdrs = [str(h).strip().lower() if h is not None else '' for h in raw_rows[0]]

    def _find(candidates):
        for kw in candidates:
            for i, rh in enumerate(raw_hdrs):
                if kw in rh:
                    return i
        return None

    ci_month      = _find(['month'])
    ci_ship_id    = _find(['shipment id'])
    ci_channel    = _find(['sales channel', 'ingestion channel store', 'channel'])
    ci_sku        = _find(['sku'])
    ci_line_name  = _find(['line item name'])
    ci_line_qty   = _find(['line item qty'])
    ci_sub_bucket = _find(['sub bucket'])
    ci_main_bucket= _find(['main bucket'])
    ci_claim_st   = _find(['claim status'])
    ci_amt        = _find(['amt'])
    ci_expected   = _find(['expected claim', 'expected'])
    ci_remark     = _find(['claims remark', 'remark'])
    ci_carrier    = _find(['carrier'])
    ci_row_status = _find(['row status'])
    ci_days       = _find(['days of order'])

    print(f"[ShipBob Excel] Column map: month={ci_month} ship={ci_ship_id} "
          f"ch={ci_channel} sub={ci_sub_bucket} amt={ci_amt} exp={ci_expected}")

    def gv(row, idx):
        if idx is None or idx >= len(row):
            return ''
        v = row[idx]
        return str(v).strip() if v is not None else ''

    BUCKET_MAP = {
        'claim raised and received':     'rec',
        'claim raised but not received': 'prog',
        'pending to claim':              'pend',
        'claim window expired':          'expired',
        'rto':                           'rto',
        'cancelled':                     'cancelled',
        'delivered':                     'delivered',
        'intransit':                     'intransit',
    }
    SKIP_BUCKETS = {'delivered', 'intransit', 'other'}

    rows          = []
    monthly       = {}
    ch_agg        = {}
    kpis          = {
        'rec':     {'count': 0, 'amt': 0.0},
        'prog':    {'count': 0, 'exp': 0.0},
        'pend':    {'count': 0, 'exp': 0.0},
        'expired': {'count': 0, 'exp': 0.0},
    }
    prog_shipments = set()
    pend_shipments = set()
    pivot1         = {}
    pivot1_months  = set()
    shipment_qty   = {}

    for raw in raw_rows[1:]:
        if not any(v is not None for v in raw):
            continue

        month        = gv(raw, ci_month)
        shipment_id  = gv(raw, ci_ship_id)
        channel      = gv(raw, ci_channel)
        sku          = gv(raw, ci_sku)
        line_name    = gv(raw, ci_line_name) or sku
        sub_bucket   = gv(raw, ci_sub_bucket)
        main_bucket  = gv(raw, ci_main_bucket)
        claim_status = gv(raw, ci_claim_st)
        remark       = gv(raw, ci_remark)
        carrier      = gv(raw, ci_carrier)

        try: amt = float(raw[ci_amt]) if ci_amt is not None and raw[ci_amt] is not None else 0.0
        except: amt = 0.0
        try: exp = float(raw[ci_expected]) if ci_expected is not None and raw[ci_expected] is not None else 0.0
        except: exp = 0.0
        try: days = float(raw[ci_days]) if ci_days is not None and raw[ci_days] is not None else 0.0
        except: days = 0.0

        bk = BUCKET_MAP.get(sub_bucket.strip().lower(),
             BUCKET_MAP.get(main_bucket.strip().lower(), 'other'))

        sb_label = (sub_bucket or main_bucket or 'Unknown').strip()
        if month and sb_label:
            if sb_label not in pivot1:
                pivot1[sb_label] = {}
            pivot1[sb_label][month] = pivot1[sb_label].get(month, 0) + 1
            pivot1_months.add(month)

        if shipment_id:
            try: qty_val = float(raw[ci_line_qty]) if ci_line_qty is not None and raw[ci_line_qty] else 0.0
            except: qty_val = 0.0
            shipment_qty[shipment_id] = shipment_qty.get(shipment_id, 0) + qty_val

        if bk in SKIP_BUCKETS:
            continue
        claim_bks = {'rec', 'prog', 'pend', 'expired'}
        if bk not in claim_bks and amt == 0 and exp == 0:
            continue

        row = {
            'month':        month,
            'shipment_id':  shipment_id,
            'channel':      channel,
            'sku':          sku,
            'line_name':    line_name,
            'sub_bucket':   sub_bucket or main_bucket,
            'bucket_key':   bk,
            'claim_status': claim_status,
            'amt':          round(amt, 2),
            'exp':          round(exp, 2),
            'remark':       remark,
            'carrier':      carrier,
            'days':         int(days) if days else 0,
        }
        rows.append(row)

        if month:
            if month not in monthly:
                monthly[month] = {'rec': 0.0, 'prog': 0.0, 'pend': 0.0, 'expired': 0.0}
            if bk == 'rec':
                monthly[month]['rec']     += amt
            elif bk == 'prog':
                monthly[month]['prog']    += exp
            elif bk == 'pend':
                monthly[month]['pend']    += exp
            elif bk == 'expired':
                monthly[month]['expired'] += exp

        if bk in kpis:
            if bk == 'rec':
                kpis[bk]['count'] += 1
                kpis[bk]['amt']   += amt
            elif bk == 'prog':
                prog_shipments.add(shipment_id)
                kpis[bk]['exp']   += exp
            elif bk == 'pend':
                pend_shipments.add(shipment_id)
                kpis[bk]['exp']   += exp
            elif bk == 'expired':
                kpis[bk]['count'] += 1
                kpis[bk]['exp']   += exp

        if channel:
            if channel not in ch_agg:
                ch_agg[channel] = {'rec': 0.0, 'prog': 0.0, 'pend': 0.0}
            if bk == 'rec':
                ch_agg[channel]['rec'] += amt
            elif bk == 'prog':
                ch_agg[channel]['prog'] += exp
            elif bk == 'pend':
                ch_agg[channel]['pend'] += exp

    kpis['prog']['count'] = len(prog_shipments)
    kpis['pend']['count'] = len(pend_shipments)

    for row in rows:
        row['total_qty'] = int(shipment_qty.get(row['shipment_id'], 0))

    months  = sorted(monthly.keys(),
                     key=lambda m: MONTH_ORDER.index(m) if m in MONTH_ORDER else 99)
    p1_months = sorted(pivot1_months,
                       key=lambda m: MONTH_ORDER.index(m) if m in MONTH_ORDER else 99)

    for m in monthly:
        for k in monthly[m]:
            monthly[m][k] = round(monthly[m][k], 2)
    for ch in ch_agg:
        for k in ch_agg[ch]:
            ch_agg[ch][k] = round(ch_agg[ch][k], 2)
    for bk in kpis:
        for k in kpis[bk]:
            if isinstance(kpis[bk][k], float):
                kpis[bk][k] = round(kpis[bk][k], 2)

    print(f"[ShipBob Excel] Loaded {len(rows)} claim rows. "
          f"rec={kpis['rec']['amt']} prog={kpis['prog']['exp']} pend={kpis['pend']['exp']}")
    return {
        'rows':          rows,
        'monthly':       monthly,
        'months':        months,
        'kpis':          kpis,
        'channels':      ch_agg,
        'pivot1':        pivot1,
        'pivot1_months': p1_months,
    }
