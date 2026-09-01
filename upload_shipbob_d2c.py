# upload_shipbob_d2c.py
# ─────────────────────
# Reads the latest ShipBob weekly Excel from C:/Users/Admin/Shibob D2C Claim/
# Filters to claim rows only (Main Bucket = "Under Dispute", ~200 rows)
# Writes them into the "ShipBob D2C Claims" tab of the recon Google Sheet.
#
# Run every Friday after dropping the new file in the folder:
#     python upload_shipbob_d2c.py

import glob
import json
import os
import re
import sys
import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────
FOLDER      = r'G:\My Drive\Shibob D2C Claim'
SHEET_ID    = '1N8qozEIZUg2FWYRqdO4UGHvcHiTtfoZr_BK_UojDSSA'   # recon sheet
TAB_NAME    = 'ShipBob D2C Claims'
CONFIG_FILE = os.path.join(os.path.dirname(__file__), 'config.json')

# ── Find latest Excel file ────────────────────────────────────────────────────
def find_latest_excel(folder):
    """Pick the most recently modified .xlsx in the folder root (not subfolders, not temp files)."""
    files = [
        f for f in glob.glob(os.path.join(folder, '*.xlsx'))
        if not os.path.basename(f).startswith('~$')   # skip Excel temp/lock files
    ]
    if not files:
        raise FileNotFoundError(f'No .xlsx files found in {folder}')
    latest = max(files, key=os.path.getmtime)
    return latest


# ── Read Excel, filter to claim rows ─────────────────────────────────────────
def extract_claim_rows(path):
    print(f'Reading: {os.path.basename(path)}')
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    if 'Dump' not in wb.sheetnames:
        raise ValueError(f"Sheet 'Dump' not found. Sheets: {wb.sheetnames}")

    ws = wb['Dump']
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    raw_headers = [str(v).strip() if v is not None else '' for v in next(rows_iter)]

    # Key columns — no PII (name/email/address excluded)
    KEEP = [
        'Month', 'Shipment ID', 'Store Order ID', 'Import Date', 'Days of order',
        'Sales Channel', 'SKU', 'Line Item Name', 'Line Item Qty',
        'Sub Bucket', 'Main Bucket', 'Claim status', 'Amt',
        'Expected Claim amount', 'Claims remark', 'Carrier',
        'Delivery Status', 'Delivery remark', 'Row Status',
    ]
    # Build index map (case-insensitive exact match)
    hi = {}
    for col in KEEP:
        for i, h in enumerate(raw_headers):
            if h.lower() == col.lower():
                hi[col] = i
                break

    missing = [c for c in KEEP if c not in hi]
    if missing:
        print(f'  Warning: columns not found (will be blank): {missing}')

    all_rows = []
    total = 0
    for raw in rows_iter:
        if not any(raw):
            continue
        total += 1
        row_vals = []
        for col in KEEP:
            if col in hi and hi[col] < len(raw):
                v = raw[hi[col]]
                if hasattr(v, 'strftime'):
                    v = v.strftime('%Y-%m-%d')
                row_vals.append(str(v) if v is not None else '')
            else:
                row_vals.append('')
        all_rows.append(row_vals)

    wb.close()
    print(f'  Total rows uploaded: {total:,}')
    return KEEP, all_rows


# ── Write to Google Sheets ────────────────────────────────────────────────────
def upload_to_sheets(headers, data_rows):
    from auth import get_sheets_credentials

    with open(CONFIG_FILE) as f:
        config = json.load(f)

    creds   = get_sheets_credentials(config.get('credentials_file'), config.get('token_file'))

    from googleapiclient.discovery import build
    service = build('sheets', 'v4', credentials=creds)

    # ── Ensure tab exists ─────────────────────────────────────────────────
    meta = service.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
    existing_tabs = [s['properties']['title'] for s in meta['sheets']]

    if TAB_NAME not in existing_tabs:
        print(f'  Creating tab: "{TAB_NAME}"')
        service.spreadsheets().batchUpdate(
            spreadsheetId=SHEET_ID,
            body={'requests': [{'addSheet': {'properties': {'title': TAB_NAME}}}]}
        ).execute()
    else:
        print(f'  Tab "{TAB_NAME}" already exists — will overwrite data')

    # ── Expand sheet dimensions to fit all rows ───────────────────────────
    sheet_meta = service.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
    sheet_id_int = next(
        s['properties']['sheetId'] for s in sheet_meta['sheets']
        if s['properties']['title'] == TAB_NAME
    )
    needed_rows = len(data_rows) + 10
    needed_cols = len(headers) + 2
    service.spreadsheets().batchUpdate(
        spreadsheetId=SHEET_ID,
        body={'requests': [{'updateSheetProperties': {
            'properties': {
                'sheetId': sheet_id_int,
                'gridProperties': {'rowCount': needed_rows, 'columnCount': needed_cols}
            },
            'fields': 'gridProperties.rowCount,gridProperties.columnCount'
        }}]}
    ).execute()
    print(f'  Sheet expanded to {needed_rows} rows × {needed_cols} cols')

    # ── Clear existing content ────────────────────────────────────────────
    service.spreadsheets().values().clear(
        spreadsheetId=SHEET_ID,
        range=f"'{TAB_NAME}'!A1:ZZ100000"
    ).execute()

    # ── Write header + data in chunks of 5 000 rows (API size limit) ─────
    CHUNK = 5000
    all_rows = [headers] + data_rows
    for i in range(0, len(all_rows), CHUNK):
        chunk = all_rows[i:i + CHUNK]
        service.spreadsheets().values().update(
            spreadsheetId=SHEET_ID,
            range=f"'{TAB_NAME}'!A{i + 1}",
            valueInputOption='RAW',
            body={'values': chunk}
        ).execute()
        print(f'  → batch {i // CHUNK + 1}: rows {i + 1}–{i + len(chunk)}')

    print(f'  ✅ Uploaded {len(data_rows)} rows + header to "{TAB_NAME}"')


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    try:
        path = find_latest_excel(FOLDER)
        headers, all_rows = extract_claim_rows(path)

        if not all_rows:
            print('No rows found — nothing uploaded.')
            sys.exit(0)

        upload_to_sheets(headers, all_rows)
        print('\nDone! Go to the dashboard and hit Refresh.')

    except Exception as e:
        print(f'\n❌ Error: {e}')
        import traceback; traceback.print_exc()
        sys.exit(1)
