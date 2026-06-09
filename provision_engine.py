import io
import base64
import openpyxl
from collections import defaultdict
from datetime import datetime
from googleapiclient.discovery import build

MONTH_NAMES = {
    1: 'January', 2: 'February', 3: 'March', 4: 'April',
    5: 'May', 6: 'June', 7: 'July', 8: 'August',
    9: 'September', 10: 'October', 11: 'November', 12: 'December'
}


def _f(val):
    if val is None:
        return 0.0
    s = str(val).strip()
    if s in ('', '#NAME?', '#REF!', '#VALUE!', '#DIV/0!', '#N/A', '#NULL!', 'None'):
        return 0.0
    try:
        return float(s.replace(',', '').replace('%', ''))
    except (ValueError, TypeError):
        return 0.0


def _s(val):
    return '' if val is None else str(val).strip()


def _clean(val):
    """Like _s but returns '' for Excel formula errors."""
    s = _s(val)
    return '' if s in ('#REF!', '#NAME?', '#VALUE!', '#DIV/0!', '#N/A', '#NULL!', 'None') else s


def _date_str(val):
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%d-%m-%Y')
    s = str(val).strip()
    for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(s[:10], fmt[:len(s[:10])]).strftime('%d-%m-%Y')
        except ValueError:
            pass
    # Try full datetime strings like "2026-04-05 13:18:36"
    try:
        return datetime.strptime(s[:19], '%Y-%m-%d %H:%M:%S').strftime('%d-%m-%Y')
    except ValueError:
        pass
    return s


def get_sheet_carriers(creds, sheet_id, month, year):
    """Return sorted list of unique carriers for the given month from Google Sheet."""
    service = build('sheets', 'v4', credentials=creds)
    result = service.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range="'Monthly Inward Logistics Report'!A1:BZ5000"
    ).execute()
    rows = result.get('values', [])

    header_idx = None
    col_map = {}
    for i, row in enumerate(rows):
        if len(row) > 2 and row[2] == 'Sale Order Number':
            header_idx = i
            for j, cell in enumerate(row):
                if cell and _s(cell) not in col_map:
                    col_map[_s(cell)] = j
            break

    if header_idx is None or 'Carrier' not in col_map:
        return []

    carrier_col = col_map['Carrier']
    date_col = col_map.get('Date')
    carriers = set()
    for row in rows[header_idx + 1:]:
        if not row or len(row) <= carrier_col:
            continue
        if date_col is not None and len(row) > date_col:
            try:
                dt = datetime.strptime(_s(row[date_col]), '%d-%m-%Y')
                if dt.month != month or dt.year != year:
                    continue
            except ValueError:
                continue
        c = _s(row[carrier_col])
        if c:
            carriers.add(c)
    return sorted(carriers)


def get_carriers_for_finance_file(creds, sheet_id, finance_bytes):
    """Extract EX numbers from finance file, look them up in Google Sheet, return unique carriers."""
    # Step 1: get EX numbers from finance file
    finance = _read_finance_data(finance_bytes)
    ex_numbers = set(finance.keys())
    if not ex_numbers:
        return []

    # Step 2: read sheet and find carriers for those EX numbers
    service = build('sheets', 'v4', credentials=creds)
    result = service.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range="'Monthly Inward Logistics Report'!A1:BZ5000"
    ).execute()
    rows = result.get('values', [])

    header_idx = None
    col_map = {}
    for i, row in enumerate(rows):
        if len(row) > 2 and row[2] == 'Sale Order Number':
            header_idx = i
            for j, cell in enumerate(row):
                if cell and _s(cell) not in col_map:
                    col_map[_s(cell)] = j
            break

    if header_idx is None or 'Carrier' not in col_map:
        return []

    so_col = col_map.get('Sale Order Number', 2)
    carrier_col = col_map['Carrier']
    carriers = set()
    for row in rows[header_idx + 1:]:
        if not row or len(row) <= max(so_col, carrier_col):
            continue
        so = _s(row[so_col])
        if so in ex_numbers:
            c = _s(row[carrier_col])
            if c:
                carriers.add(c)
    return sorted(carriers)


def _fetch_sheet_data(creds, sheet_id, month, year):
    """Fetch Monthly Inward Logistics Report, filter by month/year in Python."""
    service = build('sheets', 'v4', credentials=creds)
    result = service.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range="'Monthly Inward Logistics Report'!A1:BZ5000"
    ).execute()
    rows = result.get('values', [])

    header_idx = None
    col_map = {}
    for i, row in enumerate(rows):
        if len(row) > 2 and row[2] == 'Sale Order Number':
            header_idx = i
            for j, cell in enumerate(row):
                if cell and _s(cell) not in col_map:
                    col_map[_s(cell)] = j
            break

    if header_idx is None:
        return {}

    def gcol(row, name, default=''):
        idx = col_map.get(name)
        if idx is None or idx >= len(row):
            return default
        return row[idx]

    sheet_data = {}
    for row in rows[header_idx + 1:]:
        if not row or len(row) < 3:
            continue
        date_str = _s(gcol(row, 'Date'))
        try:
            dt = datetime.strptime(date_str, '%d-%m-%Y')
            if dt.month != month or dt.year != year:
                continue
        except ValueError:
            continue
        so = _s(gcol(row, 'Sale Order Number'))
        if not so.startswith('EX-'):
            continue
        sheet_data[so] = {
            'carrier':       _s(gcol(row, 'Carrier')),
            'mode':          _s(gcol(row, 'Mode')),
            'ior':           _s(gcol(row, 'IOR')),
            'vol_metric':    _f(gcol(row, 'Volume metric')),
            'gr_wt':         _f(gcol(row, 'Gr Wt')),
            'chargeable_wt': _f(gcol(row, 'Chargeable wt')),
            'dest_charges':  _f(gcol(row, 'Destination charges')),
        }
    return sheet_data


def _read_finance_data(excel_bytes):
    """Group Finance sheet rows (Tally export) by Sale Order Number (EX-XXXX).

    The Finance file has ~72 columns. EX-XXXX sale orders live in a column
    called 'Original Sale No' (index 46). We auto-detect which column via
    two passes: header name first, EX-XXXX pattern scan second.
    """
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)

    header_idx = None
    col_map = {}
    so_col = None
    ws = None

    # Pass 1: look for 'Sale Order Number' OR 'Original Sale No' header
    SO_HEADERS = {'Sale Order Number', 'Original Sale No', 'Order Number', 'Shipment Number'}
    for sheet_name in wb.sheetnames:
        candidate = wb[sheet_name]
        for i, row in enumerate(candidate.iter_rows(values_only=True, max_row=30)):
            if not row:
                continue
            row_vals = [_s(c) for c in row]
            found = next((h for h in SO_HEADERS if h in row_vals), None)
            if found:
                header_idx = i
                ws = candidate
                col_map = {name: k for k, name in enumerate(row_vals) if name}
                so_col = col_map[found]
                if found != 'Sale Order Number':
                    col_map['Sale Order Number'] = so_col
                break
        if ws is not None:
            break

    # Pass 2: scan ALL cells for EX-XXXX pattern
    if ws is None:
        for sheet_name in wb.sheetnames:
            candidate = wb[sheet_name]
            all_rows = list(candidate.iter_rows(values_only=True))
            ex_col_counts = defaultdict(int)
            for row in all_rows:
                if not row:
                    continue
                for j, val in enumerate(row):
                    if _s(val).startswith('EX-') and len(_s(val)) > 3:
                        ex_col_counts[j] += 1
            if not ex_col_counts:
                continue
            so_col = max(ex_col_counts, key=ex_col_counts.get)
            ws = candidate
            for i, row in enumerate(all_rows):
                if row and len(row) > so_col and _s(row[so_col]).startswith('EX-'):
                    for h in range(i - 1, -1, -1):
                        if all_rows[h] and any(_s(c) for c in all_rows[h]):
                            header_idx = h
                            col_map = {_s(c): j for j, c in enumerate(all_rows[h]) if _s(c)}
                            col_map['Sale Order Number'] = so_col
                            break
                    if header_idx is None:
                        header_idx = 0
                    break
            break

    if ws is None or so_col is None:
        debug = []
        for n in wb.sheetnames:
            first = list(wb[n].iter_rows(values_only=True, max_row=3))
            debug.append(f"Sheet '{n}': {[[_s(c) for c in r] for r in first]}")
        raise ValueError(
            f"Could not find EX-XXXX Sale Order Numbers in any sheet. "
            f"Please upload the Finance export file (Tally/sales export). "
            f"File structure: {'; '.join(debug)}"
        )

    rows = list(ws.iter_rows(values_only=True))

    def gcol(row, *names, default=None):
        for name in names:
            idx = col_map.get(name)
            if idx is not None and idx < len(row):
                val = row[idx]
                if val is not None:
                    return val
        return default

    groups = defaultdict(lambda: {
        'date': '', 'product_name': '', 'sku': '',
        'carrier': '', 'mode': '', 'ior': '', 'duty_rate': 0.0,
        'qty': 0.0, 'sales_inr': 0.0, 'duty_cost': 0.0,
    })

    start = (header_idx + 1) if header_idx is not None else 0
    for row in rows[start:]:
        if not row or len(row) <= so_col:
            continue
        so = _s(row[so_col])
        if not so.startswith('EX-'):
            continue

        g = groups[so]
        if not g['date']:
            g['date'] = _date_str(
                gcol(row, 'Date', 'Original Invoice Date', 'Invoice date',
                     'Date of Invoice', 'Dispatch Date/Cancellation Dat', default='')
            )
        if not g['product_name']:
            g['product_name'] = _clean(gcol(row, 'Product Name', 'Item', 'Description', default=''))
        if not g['sku']:
            g['sku'] = _clean(gcol(row, 'Product SKU Code', 'SKU', 'SKU Code', default=''))
        if not g['carrier']:
            g['carrier'] = _clean(gcol(row, 'Carrier', default=''))
        if not g['mode']:
            g['mode'] = _clean(gcol(row, 'Mode', default=''))
        if not g['ior']:
            g['ior'] = _clean(gcol(row, 'IOR', default=''))
        if not g['duty_rate']:
            g['duty_rate'] = _f(gcol(row, 'Duty', 'Duty %', default=None))

        g['qty'] += _f(gcol(row, 'Qty', 'Quantity', default=None))
        g['sales_inr'] += _f(gcol(row, 'Total', 'Total Amount', 'Sales Amount', default=None))
        g['duty_cost'] += _f(gcol(row, 'Duty cost', 'Duty Cost', default=None))

    return dict(groups)


def _read_logistics_data(logistics_bytes, month, year):
    """Read freight data from a Google Sheet export Excel (Monthly Inward Logistics Report).

    Looks for any sheet with a header row containing 'Sale Order Number' and 'Carrier'.
    Filters rows by the given month/year using the 'Date' column.
    """
    wb = openpyxl.load_workbook(io.BytesIO(logistics_bytes), data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        header_idx = None
        col_map = {}
        for i, row in enumerate(rows[:30]):
            row_vals = [_s(c) for c in (row or [])]
            if 'Sale Order Number' in row_vals:
                header_idx = i
                col_map = {name: j for j, name in enumerate(row_vals) if name}
                break

        if header_idx is None or 'Carrier' not in col_map:
            continue

        def gcol(row, *names, default=''):
            for name in names:
                idx = col_map.get(name)
                if idx is not None and idx < len(row):
                    val = row[idx]
                    if val is not None:
                        return val
            return default

        sheet_data = {}
        for row in rows[header_idx + 1:]:
            if not row or len(row) < 3:
                continue

            date_val = gcol(row, 'Date', 'Shipment Date')
            date_str = _date_str(date_val) if date_val else ''
            try:
                dt = datetime.strptime(date_str, '%d-%m-%Y')
                if dt.month != month or dt.year != year:
                    continue
            except ValueError:
                continue

            so = _s(gcol(row, 'Sale Order Number'))
            if not so.startswith('EX-'):
                continue

            sheet_data[so] = {
                'carrier': _s(gcol(row, 'Carrier')),
                'mode': _s(gcol(row, 'Mode')),
                'ior': _s(gcol(row, 'IOR')),
                'vol_metric': _f(gcol(row, 'Volume metric', 'Volume Metric', 'Vol Metric')),
                'gr_wt': _f(gcol(row, 'Gr Wt', 'Gross Weight', 'GR Wt')),
                'chargeable_wt': _f(gcol(row, 'Chargeable wt', 'Chargeable Weight')),
                'dest_charges': _f(gcol(row, 'Destination charges', 'Destination Charges', 'Dest Charges')),
            }

        if sheet_data:
            return sheet_data

    return {}


def process_provision(finance_bytes, month, year, creds, sheet_id, carrier_rates):
    """Main: Finance Excel + Google Sheet → rows, summary, base64 Excel."""
    finance = _read_finance_data(finance_bytes)
    sheet = _fetch_sheet_data(creds, sheet_id, month, year)
    print(f"[Provision] Sheet data: {len(sheet)} orders for {MONTH_NAMES.get(month)} {year}")

    rows = []
    unmatched = []

    for so in sorted(finance.keys(), key=lambda x: (finance[x]['date'], x)):
        fin = finance[so]
        sh = sheet.get(so)

        carrier = (sh['carrier'] if sh and sh['carrier'] else None) or fin['carrier']
        mode = (sh['mode'] if sh and sh['mode'] else None) or fin['mode']
        ior = (sh['ior'] if sh and sh['ior'] else None) or fin['ior']
        vol_metric = sh['vol_metric'] if sh else 0.0
        gr_wt = sh['gr_wt'] if sh else 0.0
        chargeable_wt = sh['chargeable_wt'] if sh else 0.0
        dest_charges = sh['dest_charges'] if sh else 0.0

        rate = _f(carrier_rates.get(carrier, 0))
        freight_cost = rate * vol_metric
        total_freight = freight_cost + dest_charges

        if not sh:
            unmatched.append(so)

        rows.append({
            'sale_order': so,
            'date': fin['date'],
            'product_name': fin['product_name'],
            'sku': fin['sku'],
            'mode': mode,
            'carrier': carrier,
            'ior': ior,
            'qty': fin['qty'],
            'sales_inr': fin['sales_inr'],
            'duty_rate': fin['duty_rate'],
            'duty_cost': fin['duty_cost'],
            'vol_metric': vol_metric,
            'gr_wt': gr_wt,
            'chargeable_wt': chargeable_wt,
            'dest_charges': dest_charges,
            'freight_cost': freight_cost,
            'total_freight': total_freight,
            'matched': bool(sh),
        })

    summary = {
        'month': MONTH_NAMES.get(month, str(month)),
        'year': year,
        'total_orders': len(rows),
        'total_qty': sum(r['qty'] for r in rows),
        'total_sales_inr': sum(r['sales_inr'] for r in rows),
        'total_duty_cost': sum(r['duty_cost'] for r in rows),
        'total_vol_metric': sum(r['vol_metric'] for r in rows),
        'total_gr_wt': sum(r['gr_wt'] for r in rows),
        'total_chargeable_wt': sum(r['chargeable_wt'] for r in rows),
        'total_dest_charges': sum(r['dest_charges'] for r in rows),
        'total_freight_cost': sum(r['freight_cost'] for r in rows),
        'total_freight_all': sum(r['total_freight'] for r in rows),
        'unmatched_orders': unmatched,
    }

    excel_out = _generate_excel(rows, summary)
    return {
        'rows': rows,
        'summary': summary,
        'excel_b64': base64.b64encode(excel_out).decode('utf-8'),
        'filename': f"{summary['month']}_{year}_Billing_Provision.xlsx",
    }


def _generate_excel(rows, summary):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    DARK = "1F3864"
    ALT = "EBF3FB"
    wb = openpyxl.Workbook()

    # ── Sheet 1: Pivot ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = f"{summary['month'][:3]} {summary['year']} Provision"

    headers = [
        'Sale Order Number', 'Date', 'Product Name', 'Product SKU',
        'Mode', 'Carrier', 'IOR', 'Qty', 'Sales (₹)',
        'Duty %', 'Duty Cost (₹)', 'Volume Metric (kg)', 'Gr Wt (kg)',
        'Chargeable Wt (kg)', 'Destination Charges (₹)',
        'Freight Cost (₹)', 'Total Freight Cost (₹)',
    ]
    nc = len(headers)

    ws.merge_cells(f'A1:{get_column_letter(nc)}1')
    t = ws['A1']
    t.value = (
        f"Consolidated by Sale Order Number  |  "
        f"{summary['month'].upper()} {summary['year']} — MONTHLY BILLING PROVISION"
    )
    t.font = Font(bold=True, color="FFFFFF", size=11)
    t.fill = PatternFill("solid", fgColor=DARK)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    for j, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 32

    for i, row in enumerate(rows):
        r = i + 3
        fill = PatternFill("solid", fgColor=ALT) if i % 2 == 0 else None
        vals = [
            row['sale_order'], row['date'], row['product_name'], row['sku'],
            row['mode'], row['carrier'], row['ior'],
            row['qty'], row['sales_inr'],
            f"{row['duty_rate'] * 100:.1f}%" if row['duty_rate'] else '',
            row['duty_cost'], row['vol_metric'], row['gr_wt'],
            row['chargeable_wt'], row['dest_charges'],
            row['freight_cost'], row['total_freight'],
        ]
        for j, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=j, value=val)
            if fill:
                c.fill = fill
            if j >= 8 and j != 10:
                c.number_format = '#,##0.00'
            c.alignment = Alignment(
                horizontal='right' if j >= 8 else 'left',
                vertical='center'
            )

    tr = len(rows) + 3
    total_map = {
        8: summary['total_qty'],
        9: summary['total_sales_inr'],
        11: summary['total_duty_cost'],
        12: summary['total_vol_metric'],
        13: summary['total_gr_wt'],
        14: summary['total_chargeable_wt'],
        15: summary['total_dest_charges'],
        16: summary['total_freight_cost'],
        17: summary['total_freight_all'],
    }
    for j in range(1, nc + 1):
        c = ws.cell(row=tr, column=j, value='GRAND TOTAL' if j == 1 else total_map.get(j))
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=DARK)
        if j in total_map:
            c.number_format = '#,##0.00'
        c.alignment = Alignment(
            horizontal='center' if j == 1 else 'right',
            vertical='center'
        )

    ws.freeze_panes = 'A3'
    for j, w in enumerate([20, 12, 35, 28, 14, 20, 8, 10, 15, 8, 14, 16, 11, 15, 20, 15, 18], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ── Sheet 2: Summary ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.merge_cells('A1:B1')
    t2 = ws2['A1']
    t2.value = f"{summary['month'].upper()} {summary['year']} — BILLING PROVISION SUMMARY"
    t2.font = Font(bold=True, color="FFFFFF", size=13)
    t2.fill = PatternFill("solid", fgColor=DARK)
    t2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    for j, label in [(1, 'Metric'), (2, 'Value')]:
        c = ws2.cell(row=2, column=j, value=label)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E4057")
        c.alignment = Alignment(horizontal='center')

    metrics = [
        ('Total Sale Orders', summary['total_orders']),
        ('Total Qty Shipped', summary['total_qty']),
        ('Total Sales Value (₹)', summary['total_sales_inr']),
        ('Total Duty Cost (₹)', summary['total_duty_cost']),
        ('Total Volume Metric', summary['total_vol_metric']),
        ('Total Gross Weight (kg)', summary['total_gr_wt']),
        ('Total Chargeable Weight (kg)', summary['total_chargeable_wt']),
        ('Total Destination Charges (₹)', summary['total_dest_charges']),
        ('Total Freight Cost (₹)', summary['total_freight_cost']),
        ('Total Freight Cost incl. all (₹)', summary['total_freight_all']),
    ]
    for i, (metric, val) in enumerate(metrics, 3):
        c1 = ws2.cell(row=i, column=1, value=metric)
        c2 = ws2.cell(row=i, column=2, value=val)
        c2.number_format = '#,##0.00'
        c1.alignment = Alignment(horizontal='left', vertical='center')
        c2.alignment = Alignment(horizontal='right', vertical='center')
        if i % 2 == 0:
            for col in [1, 2]:
                ws2.cell(row=i, column=col).fill = PatternFill("solid", fgColor=ALT)

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 22

    if summary['unmatched_orders']:
        r = len(metrics) + 4
        c = ws2.cell(row=r, column=1,
                     value=f"Unmatched Orders ({len(summary['unmatched_orders'])}) — no Sheet data:")
        c.font = Font(bold=True, color="FF0000")
        ws2.cell(row=r + 1, column=1, value=', '.join(summary['unmatched_orders']))

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
