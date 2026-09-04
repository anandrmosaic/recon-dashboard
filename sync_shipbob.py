"""
sync_shipbob.py — Push latest ShipBob Excel → "ShipBob D2C Claims" Sheets tab

Run every Thursday after saving the new weekly Excel to:
  G:\My Drive\Shibob D2C Claim\Shipbob Order Data - Updated DD-MM-YYYY.xlsx

Usage:
  python sync_shipbob.py
"""
import json, os, glob, sys
import openpyxl
sys.path.insert(0, r'C:\Users\Admin\recon-dashboard')
from auth import get_sheets_credentials
from googleapiclient.discovery import build

with open(r'C:\Users\Admin\recon-dashboard\config.json') as f:
    cfg = json.load(f)

FOLDER   = cfg.get('shipbob_d2c_folder', r'G:\My Drive\Shibob D2C Claim')
SHEET_ID = cfg['recon_sheet_id']
TAB_NAME = cfg.get('shipbob_d2c_tab', 'ShipBob D2C Claims')

# ── Find latest Excel ─────────────────────────────────────────────────────────
pattern = os.path.join(FOLDER, "Shipbob Order Data - Updated *.xlsx")
files   = sorted(glob.glob(pattern), key=os.path.getmtime)
if not files:
    print(f"❌  No Excel file found in: {FOLDER}")
    sys.exit(1)

latest = files[-1]
print(f"📂  Reading: {os.path.basename(latest)}")

wb = openpyxl.load_workbook(latest, read_only=True, data_only=True)
ws = wb["Dump"] if "Dump" in wb.sheetnames else wb.active
print(f"📄  Sheet: {ws.title}")

# Read all rows (convert to strings for Sheets API)
all_rows = []
for row in ws.iter_rows(values_only=True):
    str_row = []
    for v in row:
        if v is None:
            str_row.append('')
        elif isinstance(v, float) and v == int(v):
            str_row.append(str(int(v)))
        else:
            str_row.append(str(v))
    all_rows.append(str_row)
wb.close()

print(f"📊  Rows to sync: {len(all_rows):,}")

# ── Connect to Sheets ─────────────────────────────────────────────────────────
creds   = get_sheets_credentials(cfg['credentials_file'], cfg['token_file'])
service = build('sheets', 'v4', credentials=creds)
ss      = service.spreadsheets()

# Check if tab exists; create it if not
meta      = ss.get(spreadsheetId=SHEET_ID).execute()
tab_names = [s['properties']['title'] for s in meta['sheets']]

if TAB_NAME not in tab_names:
    print(f"➕  Creating tab '{TAB_NAME}'…")
    ss.batchUpdate(
        spreadsheetId=SHEET_ID,
        body={'requests': [{'addSheet': {'properties': {'title': TAB_NAME}}}]}
    ).execute()
    print(f"✅  Tab created.")
else:
    print(f"🗑️   Clearing existing '{TAB_NAME}' tab…")
    ss.values().clear(
        spreadsheetId=SHEET_ID,
        range=f"'{TAB_NAME}'!A1:BZ100000"
    ).execute()

# ── Write data in batches of 5000 rows ───────────────────────────────────────
BATCH = 5000
total = len(all_rows)
written = 0

for start in range(0, total, BATCH):
    chunk    = all_rows[start:start + BATCH]
    end_row  = start + len(chunk)
    range_a1 = f"'{TAB_NAME}'!A{start + 1}"
    ss.values().update(
        spreadsheetId=SHEET_ID,
        range=range_a1,
        valueInputOption='RAW',
        body={'values': chunk}
    ).execute()
    written += len(chunk)
    print(f"  ✍️   Written {written:,} / {total:,} rows…")

print(f"\n✅  Sync complete — {written:,} rows pushed to '{TAB_NAME}'")
print(f"    Hit Refresh on the dashboard to see updated numbers.")
